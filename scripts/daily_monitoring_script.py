import ctypes
ctypes.windll.kernel32.SetThreadExecutionState(0x80000002)

while True:

    import os
    from bs4 import BeautifulSoup
    import pandas as pd
    import inflect
    import time
    from datetime import date
    from datetime import timedelta 
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Color, PatternFill, Font, Border, Side
    from openpyxl.styles import colors
    from openpyxl.cell import Cell
    import urllib.request as urllib2
    from openpyxl.utils.dataframe import dataframe_to_rows
    import win32com.client as win32
    from openpyxl.worksheet.hyperlink import Hyperlink
    import numpy as np
    import datetime
    import threading
    import sys
    import quipclient as quip
    import pytz
    import pathlib
    import tempfile
    import warnings
    import requests
    import subprocess
    from http import cookiejar
    from requests_kerberos import HTTPKerberosAuth, OPTIONAL

    warnings.filterwarnings("ignore", category=requests.packages.urllib3.exceptions.InsecureRequestWarning)

    class MidwayAuthenticator:
        def __init__(self):
            self.session = requests.Session()
            self.kerberos_auth = HTTPKerberosAuth(mutual_authentication=OPTIONAL)
        
        def _get_midway_cookies(self):
            home_folder = pathlib.Path.home()
            midway_cookie_file = os.path.join(home_folder, ".midway", "cookie")
            
            if not os.path.exists(midway_cookie_file):
                print("Cookie file not found. Running mwinit...")
                subprocess.run(['mwinit', '--no-aea'])
            
            with tempfile.NamedTemporaryFile(mode="w", delete=False) as temp_file:
                with open(midway_cookie_file) as midway_file:
                    for line in midway_file:
                        if line.startswith("#HttpOnly_"):
                            temp_file.write(line[10:])
                        else:
                            temp_file.write(line)
                temp_file.flush()
                
                cookies = cookiejar.MozillaCookieJar(temp_file.name)
                cookies.load(ignore_discard=True, ignore_expires=True)
            
            os.remove(temp_file.name)
            return cookies
        
        def make_request(self, url, method="get", **kwargs):
            try:
                # Load fresh cookies for each request
                cookies = self._get_midway_cookies()
                response = self.session.request(
                    method=method.lower(),
                    url=url,
                    auth=self.kerberos_auth,
                    cookies=cookies,
                    verify=False,
                    **kwargs
                )
                return response
            except Exception as e:
                print(f"Request failed: {str(e)}")
                return None

    midway_auth = MidwayAuthenticator()

    def make_request(url, method="get", **kwargs):
        return midway_auth.make_request(url, method, **kwargs)

    def verify_midway_auth(url):
        global midway_auth  # Access the global authenticator instance
        
        while True:
            response = make_request(url)
            conn_code = response.status_code
            print(f"Status code: {conn_code}")
            
            if conn_code == 200:
                print("Midway authenticated successfully!")
                return response
            else:
                print("Authentication failed. Running mwinit...")
                subprocess.run(['mwinit', '--no-aea'])
                # Create a new authenticator instance with fresh session and cookies
                midway_auth = MidwayAuthenticator()
                print("Retrying authentication...")
    login = os.getlogin()
    print(login)
    
    URL = "https://phonetool.amazon.com/users/search?query={name}"
    url = URL.format(name=login)

    response = verify_midway_auth(url)
    print(response.json())

    pdt = pytz.timezone('America/Los_Angeles')
    ist = pytz.timezone('Asia/Kolkata')
        # ES_CONTINUOUS | ES_SYSTEM_REQUIRED

    # Get today's date in YYYY-MM-DD format
    today_date = datetime.datetime.fromtimestamp(time.time()).strftime("%Y-%m-%d")
    count_file_name = f"{today_date}_runcount.txt"

    # Check if the count file for today exists
    if os.path.exists(count_file_name):
        # Read the current count and increment it
        with open(count_file_name, "r") as count_file:
            run_count = int(count_file.read().strip())
            run_count += 1
    else:
        # First run today, start count at 1
        run_count = 1

    

    # Log the current run information
    print(f"Run number for {today_date}: {run_count}")
    take = run_count
    p = inflect.engine()
    run = p.ordinal(take)
    print(run)

    user_input = None

    def get_user_input():
        global user_input
        user_input = input("Send mail to 1. Team, 2. Self: ")

    # Start the input thread
    input_thread = threading.Thread(target=get_user_input)
    input_thread.daemon = True  # Allows program to exit even if thread is running
    input_thread.start()

    # Wait for the input for 10 seconds
    input_thread.join(timeout=10)

    # Check if user input was provided; if not, set default value
    if user_input is None:
        print("\nNo input provided, defaulting to 1.")
        variable = '1'  # Default value
    # Handle the user input
    elif user_input in ['1', '2']:
        variable = user_input  # Set your desired variable here based on input
        print(f"\nYou entered: {variable}")
    else:
        print("\nInvalid input, defaulting to 1.")
        variable = '1'  # Set variable to default if invalid input

    teamss = variable
    
    #https://googlechromelabs.github.io/chrome-for-testing/
    #https://developer.chrome.com/docs/chromedriver/

    #COMMENT FROM HERE TO USE LOCALLY STORED FILES FOR WEB SCRAPPING: START
    baseurl = "https://platform.quip-amazon.com"
    #https://quip-amazon.com/dev/token
    access_token = "UFRGOU1BWEZlQTQ=|1763527256|SQHi+b9hge2suD1/dY40J82nzcID4L1+iDm7TIVsYjA="
    thread_id = 'iKqHAWP6VJ0l'
    client = quip.QuipClient(access_token = access_token, base_url=baseurl)
    rawdictionary = client.get_thread(thread_id)
    dfs=pd.read_html(rawdictionary['html'])
    df_links_core = dfs[0]
    sla_df = dfs[1]
    df_t12 = dfs[2]
    df_allclusters = dfs[3]
    df_force_jobs = dfs[4]
    df_selenium_xpaths = dfs[5]

    print(dfs)
    raw_df = dfs[-1]
    raw_df.columns=raw_df.iloc[0]
    raw_df=raw_df.iloc[1:,1:] 
    raw_df=raw_df.replace('\u200b', np.nan) 
    raw_df = raw_df.dropna(how='all')
    raw_df = raw_df.dropna(axis=1,how='all')
    print(raw_df)
    print(raw_df.info())
    print("Printing Sheet 1\n")
    print(df_links_core)
    print("Printing Sheet 2\n")
    print(sla_df)
    df_links_core.columns=df_links_core.iloc[0] 
    df_links_core=df_links_core.iloc[1:,1:] 
    df_links_core=df_links_core.replace('\u200b', np.nan) 
    df_links_core = df_links_core.dropna(how='all')
    df_links_core = df_links_core.dropna(axis=1,how='all')
    df_links = df_links_core[['link','name']].drop_duplicates()
    df_links_priority = df_links_core[['link','name','BusinessPriority','ManualPriority']].drop_duplicates()
    df_links_priority.rename(columns={'link':'clink'},inplace=True)
    df_links_priority.rename(columns={'name':'tablename'},inplace=True)
    df_links_priority.rename(columns={'BusinessPriority':'bp'},inplace=True)
    df_links_priority.rename(columns={'ManualPriority':'mp'},inplace=True)
    sla_df.columns=sla_df.iloc[0] 
    sla_df=sla_df.iloc[1:,1:] 
    sla_df=sla_df.replace('\u200b', np.nan) 
    sla_df = sla_df.dropna(how='all')
    sla_df = sla_df.dropna(axis=1,how='all')
    sla_df.rename(columns={'Benchmark Data Availability (Hrs)':'SLA'},inplace=True)
    sla_df['Table'] = sla_df['Schema'].str.upper() +"."+ sla_df['Table'].str.upper() 
    sla_df = sla_df[['Table','SLA']].drop_duplicates()

    # New code for t12
    df_t12.columns=df_t12.iloc[0] 
    df_t12=df_t12.iloc[1:,1:] 
    df_t12=df_t12.replace('\u200b', np.nan) 
    df_t12 = df_t12.dropna(how='all')
    df_t12 = df_t12.dropna(axis=1,how='all')
    df_t12.rename(columns={'FULL_TABLE_NAME':'Table'},inplace=True)
    df_t12 = df_t12[['Table']].drop_duplicates()

    df_allclusters.columns=df_allclusters.iloc[0] 
    df_allclusters=df_allclusters.iloc[1:,1:] 
    df_allclusters=df_allclusters.replace('\u200b', np.nan) 
    df_allclusters = df_allclusters.dropna(how='all')
    df_allclusters = df_allclusters.dropna(axis=1,how='all')
    df_allclusters.rename(columns={'FULL_TABLE_NAME':'Table'},inplace=True)
    df_allclusters = df_allclusters[['Table']].drop_duplicates()

    df_force_jobs.columns=df_force_jobs.iloc[0] 
    df_force_jobs=df_force_jobs.iloc[1:,1:] 
    df_force_jobs=df_force_jobs.replace('\u200b', np.nan) 
    df_force_jobs = df_force_jobs.dropna(how='all')
    df_force_jobs = df_force_jobs.dropna(axis=1,how='all')
    df_force_jobs.rename(columns={'FULL_TABLE_NAME':'Table'},inplace=True)
    df_force_jobs = df_force_jobs[['Table']].drop_duplicates()


    df_selenium_xpaths.columns=df_selenium_xpaths.iloc[0] 
    df_selenium_xpaths=df_selenium_xpaths.iloc[1:,1:] 
    df_selenium_xpaths=df_selenium_xpaths.replace('\u200b', np.nan) 
    df_selenium_xpaths = df_selenium_xpaths.dropna(how='all')
    df_selenium_xpaths = df_selenium_xpaths.dropna(axis=1,how='all')
    df_selenium_xpaths = df_selenium_xpaths.drop_duplicates()
    print("\nPrinting the force success and restart XPaths\n")
    print(df_selenium_xpaths)
    fs_text = df_selenium_xpaths['fs_text'].iloc[0]
    fs_submit = df_selenium_xpaths['fs_submit'].iloc[0]
    fs_response = df_selenium_xpaths['fs_response'].iloc[0]
    rs_text = df_selenium_xpaths['rs_text'].iloc[0]
    rs_submit = df_selenium_xpaths['rs_submit'].iloc[0]
    rs_response = df_selenium_xpaths['rs_response'].iloc[0]

    print("\nPrinting singular XPaths\n")
    print(f"fs_text:{fs_text}\n")
    print(f"fs_submit:{fs_submit}\n")
    print(f"fs_response:{fs_response}\n")
    print(f"rs_text:{rs_text}\n")
    print(f"rs_submit:{rs_submit}\n")
    print(f"rs_response:{rs_response}\n")

    print("Printing main link Table \n")
    print(df_links)
    print(df_links.info())
    print("Printing priority list Table \n")
    print(df_links_priority)
    print(df_links_priority.info())
    print("Printing SLA Table \n")
    print(sla_df)
    print(sla_df.info())
    df_links.to_excel("links.xlsx")
    df_links_priority.to_excel("linkspriority.xlsx")
    sla_df.to_excel("sla_april.xlsx")


    print("Printing T+12 required Tables \n")
    print(df_t12)
    print(df_t12.info())
    df_t12.to_excel("df_t12.xlsx")

    print("Printing all cluster required Tables \n")
    print(df_allclusters)
    print(df_allclusters.info())
    df_allclusters.to_excel("df_allclusters.xlsx")

    print("Printing force success required Tables \n")
    print(df_force_jobs)
    print(df_force_jobs.info())
    df_force_jobs.to_excel("df_force_jobs.xlsx")
    #COMMENT TILL HERE TO USE LOCALLY STORED FILES FOR WEB SCRAPPING: END
    df_links = pd.read_excel("links.xlsx")
    df_links_priority= pd.read_excel("linkspriority.xlsx")
    sla_df = pd.read_excel("sla_april.xlsx")
    df_t12 = pd.read_excel("df_t12.xlsx")
    df_allclusters = pd.read_excel("df_allclusters.xlsx")
    df_force_jobs = pd.read_excel("df_force_jobs.xlsx")
    URL = "https://phonetool.amazon.com/users/search?query={name}"
    url = URL.format(name=login)
    #url = "https://datacentral.a2z.com/dw-platform/servlet/dwp/template/DWPJobPerformanceHistory.vm/job_id/26339173"
    response = make_request(url)
    data = response.json()
    full_id = data[0]['id']

    mailname = full_id.split(' (')[0]
    print(mailname)


    current_timestamp1 = time.time()
    starttime = current_timestamp1
    dt_object1 = datetime.datetime.fromtimestamp(starttime)
    print("Start time")
    print(dt_object1)
    formatted1 = dt_object1.strftime("%Y-%m-%d %H:%M:%S")
    print(formatted1)

    filedate = date.today().strftime("%Y-%m-%d")
    start_date = (date.today() - timedelta(days = 1)).strftime("%Y-%m-%d")
    end_date = (date.today() - timedelta(days = 7)).strftime("%Y-%m-%d")
    print(start_date)
    print(end_date)


    team = 0

    if team==0:
        a="linkspriority"
    else:
        a="linkspriority"
    if team==0:
        b="links"
    else:
        b="links"

    prt = pd.read_excel(str(a)+".xlsx")
    mt = pd.read_excel(str(b)+".xlsx")

    prt  = prt.drop_duplicates()
    mt  = mt.drop_duplicates()
    df1 = mt

    print('Priority table')
    print(prt)
    print(prt.dtypes)
    print('Main table')
    print(df1)
    print(df1.dtypes)




    forcesuccess_list = ['DIGITS.F_CUSTOMER_SERVICE_3P_MMT']
    nonBOM2_list = ['DIGITS.DIM_CAMPAIGN','DIGITS.DIM_CAMPAIGN_ADVERTISERS','DIGITS.DIM_CAMPAIGN_FUNDER','DIGITS.DIM_LPA_OFFLINE_MERCHANTS','DIGITS.F_CUSTOMER_TTM_EXT','DIGITS.F_LOWER_FUNNEL','DIGITS.TH_DATA_ASIN','DIGITS.TH_DATA_ORDERID','DIGITS.TH_DATA_TRTYPE','SANDBOX.EUCAPP_FACT','SANDBOX.F_CUSTOMER_PROFILE','SANDBOX.F_CUSTOMER_TTM'] #AdHoc Task
    print(forcesuccess_list)
    print(nonBOM2_list)
    df_force_jobs = pd.read_excel('df_force_jobs.xlsx')
    df_allclusters = pd.read_excel('df_allclusters.xlsx')
    print("Reprinting")
    # Convert the 'TableNames' column to a tuple
    nonBOM2_list = tuple(df_allclusters['Table'])
    forcesuccess_list = tuple(df_force_jobs['Table'])

    print(forcesuccess_list)
    print(nonBOM2_list)



    #start test
    i=0
    df1_app = pd.DataFrame()
    loop1_exc = pd.DataFrame()
    loop1_exc2 = pd.DataFrame()
    df1_success = pd.DataFrame()
    #Enter FOR loop1
    for index, row in df1.iterrows():
        try:
            name = row['name']
            link_h = row['link']
            i+=1
            print(i)
            # Navigating to the webpage via the link
            response = make_request(link_h)
            soup = BeautifulSoup(response.content, 'html.parser')
            # Find the table you're interested in
            table = soup.find('table', class_='tablesorter')
            # Extract the table data using pandas
            table_data = pd.read_html(str(table))
            # Display the table data
            table_rows = table_data[0]

            # Getting the table
            table = soup.find('table', class_='tablesorter')
            table_data = pd.read_html(str(table))
            table_rows = table_data[0]
            print("\nPrinting top 5 rows\n")
            print(table_rows.head(5))
            # Fetching T-1 success jobruns on ideaprod_pmnt2 (No further operations except appending and sending as an attachment on mail)
            table_success = table_rows[table_rows['Dataset Date'] == start_date]
            #table_success = table_success[(table_success['Database'] == "ideaprod_pmnt2") | table_success['Database'].isnull() | table_success['Database'].isna()]
            if not table_success.empty:
                table_success['Name'] = name
                table_success = table_success[(table_success['Database'] == "ideaprod_pmnt2") | table_success['Database'].isnull() | table_success['Database'].isna() | table_success['Name'].str.startswith(tuple(nonBOM2_list))]
                table_success = table_success[table_success['Status'].str.upper() == "SUCCESS"]
                table_success['JobHistoryLink'] = link_h
                table_success['JobRunLink'] = "https://datacentral.a2z.com/console?action=jobrun_details&jobrun_id=" + table_success['Job Run'].astype(str)
                table_success['Sl No'] = i
                table_success = table_success[['Sl No','Name','Dataset Date','Schedule Hour (Local)','Status','JobRunLink','JobHistoryLink','End Date (PST)']]

                df1_success_f = pd.concat([table_success],ignore_index=True)
                df1_success = df1_success._append(df1_success_f,ignore_index=True)
                print(df1_success)
            # Fetching T-1 to T-7 non-success jobruns (All the further operations are on these data)
            table_rows = table_rows[(table_rows['Status'].str.upper() != "SUCCESS")]
            table_rows = table_rows[(table_rows['Dataset Date']>=end_date) & (table_rows['Dataset Date']<=start_date)]
            #table_rows = table_rows[(table_rows['Database'] == "ideaprod_pmnt2") | table_rows['Database'].isnull() | table_rows['Database'].isna()]
            if not table_rows.empty:
                table_rows['Name'] = name
                table_rows = table_rows[(table_rows['Database'] == "ideaprod_pmnt2") | table_rows['Database'].isnull() | table_rows['Database'].isna() | table_rows['Name'].str.startswith(tuple(nonBOM2_list))]
                table_rows['JobHistoryLink'] = link_h
                table_rows['JobRunLink'] = table_rows.apply(
                    lambda row: 
                        "https://datacentral.a2z.com/console?action=jobrun_details&jobrun_id=" + str(row['Job Run']) + "&refersh=true"
                        if row['Status'].upper() != 'ERROR' else "https://datanet-service.amazon.com/jobRunError/"+str(row['Job Run']), 
                    axis=1
                )
                print("Loop 1 link details")
                print(table_rows['JobRunLink'])
                table_rows['JobRunLink2'] = "https://datacentral.a2z.com/console?action=jobrun_details&jobrun_id=" + table_rows['Job Run'].astype(str) + "#dependencies"
                #table_rows4['flag'] = table_rows4['Table'].str.upper().str.startswith(('DIGITS', 'SANDBOX'))
                table_rows['Sl No'] = i
                table_rows = table_rows[['Sl No','Name','Dataset Date','Status','JobRunLink','JobHistoryLink','Schedule Hour (Local)','JobRunLink2']]
            print(table_rows)
            for index, row in table_rows.iterrows():
                try:
                    table_rows2 = pd.DataFrame()
                    name2 = row['Name']
                    link_jr = row['JobRunLink']
                    print("Loop 2 ke andar browse")
                    print(link_jr)
                    link_jr2 = row['JobRunLink2']
                    link_jh = row['JobHistoryLink']
                    dsdate = row['Dataset Date']
                    stat = row['Status']
                    schedule = row['Schedule Hour (Local)']
                    response2 = make_request(link_jr)
                    soup2 = BeautifulSoup(response2.content, 'html.parser')
                    if (stat.upper() == "ERROR"):
                        # table = soup2.find('table', attrs={'id': 'eventTable'})
                        # if table:
                        #     # Find the first row in the table body
                        #     row = table.find('tbody').find('tr')

                        #     # Extract the content of the event and event reason columns
                        #     event = row.find_all('td')[2].get_text(strip=True)
                        #     reason = row.find_all('td')[4].get_text(strip=True)

                        #     # Create the error message string in the desired format
                        #     error_message = f"EVENT: {event}, REASON: {reason}"
                        #     print(error_message)
                        #     details = error_message
                        # else:
                        #     print("Table with id 'eventTable' not found")
                        #     details = "Some exception occured, check mannualy"
                        data = response2.json()
                        full_message = data['message']
                        details = full_message
                    else:
                        summary_div = soup2.find('div', {'id': 'summary'})
                        paragraphs = summary_div.find_all('p')
                        print(len(paragraphs))
                        print(paragraphs)
                        if "Waiting for the following files" in str(paragraphs):
                            last_paragraph = paragraphs[1].get_text()
                        else:
                            last_paragraph = paragraphs[-1].get_text()
                        details = last_paragraph
                    print(details)
                    print(name2+"-"+stat+"-"+dsdate+"-"+"-"+details)
                    if (stat.upper() == "ERROR"):
                        response2 = make_request(link_jr2)
                        soup2 = BeautifulSoup(response2.content, 'html.parser')
                    table_dependencies_heading_file = soup2.find('h4', text='Data File Dependencies')
                    table_dependencies_heading_table = soup2.find('h4', text='Table Load Dependencies')
                    if table_dependencies_heading_file:
                        table = table_dependencies_heading_file.find_next('table')  # Find the table after the heading
                        df = pd.read_html(str(table))[0]  # This directly extracts the table into a dataframe
                        print(df)
                    elif table_dependencies_heading_table:
                        table = table_dependencies_heading_table.find_next('table')  # Find the table after the heading
                        df = pd.read_html(str(table))[0]  # This directly extracts the table into a dataframe
                        print(df)
                    else:
                        print("Data File Dependencies and Table Load Dependencies sections not found")
                    table_rowsn = df
                    print(table_rowsn)
                    print(table_rowsn.dtypes)
                    total_rows = table_rowsn.shape[0]
                    true_count = table_rowsn['Valid?'].sum()
                    unsat = total_rows - true_count
                    first_column_name = table_rowsn.columns[0]
                    print(str(total_rows)+str(true_count)+str(unsat)+str(true_count < total_rows))
                    if true_count < total_rows:
                        truth = "TRUE"
                        if first_column_name.upper() == 'DB':
                            table_rowsn = table_rowsn[table_rowsn['Valid?'] == False].reset_index()
                            def mskds(row):
                                if row['Table'].upper().startswith("DIGITS"):
                                    return 3
                                elif row['Table'].upper().startswith("SANDBOX"):
                                    return 2
                                else:
                                    return 1
                            table_rowsn['D_S'] = table_rowsn.apply(mskds,axis=1)
                            table_rowsn = table_rowsn.sort_values(by=['D_S'], ascending=[True]).reset_index()
                            table_rowsn['serial_number'] = table_rowsn.index + 1
                            table_rowsn['output'] = table_rowsn.apply(lambda row: f"{row['serial_number']}. {row['Table']} ({row['Dataset Date']})", axis=1)
                            output_row = ', '.join(table_rowsn['output'])    
                        elif first_column_name.upper() == 'FILENAME':
                            table_rowsn = table_rowsn[table_rowsn['Valid?'] == False].reset_index()
                            table_rowsn['serial_number'] = table_rowsn.index + 1
                            table_rowsn['output'] = table_rowsn.apply(lambda row: f"{row['serial_number']}. {row['Filename']}", axis=1)
                            output_row = ', '.join(table_rowsn['output'])
                    else:
                        truth = "FALSE"
                        output_row = details
                    print(output_row)
                    dependency = str(true_count)+" / "+str(total_rows)
                    print(dependency)    
                    # print(details)
                    table_rows2['Sl No'] = pd.Series([i] * 1)
                    table_rows2['Name'] = pd.Series([name2] * 1)
                    table_rows2['JobHistoryLink'] = pd.Series([link_jh] * 1)
                    table_rows2['JobRunLink'] = pd.Series([link_jr] * 1)
                    table_rows2['Dataset Date'] = pd.Series([dsdate] * 1)
                    table_rows2['Status'] = pd.Series([stat] * 1)
                    table_rows2['Details'] = pd.Series([details] * 1)
                    table_rows2['Schedule Hour (Local)'] = pd.Series([schedule] * 1)
                    table_rows2['Sat_Dependency'] = pd.Series([dependency] * 1)
                    table_rows2['WFD'] = pd.Series([output_row] * 1)
                    table_rows2['TR'] = pd.Series([truth] * 1)
                    table_rows2 = table_rows2[['Sl No','Name','Dataset Date','Status','JobRunLink','JobHistoryLink','Details','Schedule Hour (Local)','Sat_Dependency','WFD','TR']]
                    print(table_rows2)
                    print(table_rows2.dtypes)
                    df1_app_f = pd.concat([table_rows2],ignore_index=True)
                    df1_app = df1_app._append(df1_app_f,ignore_index=True)
                    print(df1_app)
                except Exception as e:
                    exc2 = pd.DataFrame()
                    name2 = row['Name']
                    link_jr = row['JobRunLink']
                    link_jh = row['JobHistoryLink']
                    dsdate = row['Dataset Date']
                    stat = row['Status']
                    schedule = row['Schedule Hour (Local)']
                    exc2['Table Name'] = pd.Series([name] * 1)
                    exc2['Link'] = pd.Series([link_h] * 1)
                    exc2['Table Name2'] = pd.Series([name2] * 1)
                    exc2['Link2'] = pd.Series([link_jh] * 1)
                    exc2['Link3'] = pd.Series([link_jr] * 1)
                    exc2['Dataset Date'] = pd.Series([dsdate] * 1)
                    exc2['Status'] = pd.Series([stat] * 1)
                    exc2['Schedule'] = pd.Series([schedule] * 1)  
                    exc2['Exception_Details'] =  pd.Series([str(e)] * 1) 

                    print(exc2)
                    loop1_exc2_f = pd.concat([exc2],ignore_index=True)
                    loop1_exc2 = loop1_exc2._append(loop1_exc2_f,ignore_index=True)

            
        except Exception as d:
            exc = pd.DataFrame()
            exc['Table Name'] = pd.Series([name] * 1)
            exc['Link'] = pd.Series([link_h] * 1)
            exc['Exception_Details'] =  pd.Series([str(d)] * 1) 
            print(exc)
            loop1_exc_f = pd.concat([exc],ignore_index=True)
            loop1_exc = loop1_exc._append(loop1_exc_f,ignore_index=True)
    #Came out of FOR loop1
    print("Printing FOR loop 2 & 1 Resultant Table")    
    print(df1_app)
    print("Printing FOR loop 1 Exception Table")    
    print(loop1_exc)
    print("Printing FOR loop 2(inside 1) Exception Table")    
    print(loop1_exc2)
    print("Printing FOR loop 1 Success Table")    
    print(df1_success)
    #End Loop 1
    df1_app.to_excel("df1_app.xlsx",index=False)
    loop1_exc.to_excel("loop1_exc.xlsx",index=False)
    loop1_exc2.to_excel("loop1_exc2.xlsx",index=False)
    df1_success.to_excel("df1_success.xlsx",index=False)
    #end test


    df1_app = pd.read_excel("df1_app.xlsx")
    df1_success = pd.read_excel("df1_success.xlsx")

    def msk(row):
        if row['Status'].upper() == "EXECUTING":
            return "Executing", row['Details'], 2
        elif row['Status'].upper() == "WAITING FOR REQUIREMENTS":
            if "JOB RUN IS IN QUEUE FOR RESOURCES" in row['Details'].upper():
                return "Waiting for Resources",row['Details'], 3
            else:
                if ("JOB RUN WILL NOT EXECUTE BEFORE" in row['Details'].upper()) & (row['TR'] == False):
                    row['det'] = row['Details'].upper().find("JOB RUN WILL NOT EXECUTE BEFORE")
                    print(row['det'])
                    details = (row['Details'][row['det']:])
                    return "Waiting for Dependencies", details, 4
                else:    
                    return "Waiting for Dependencies", row['WFD'], 4
        elif row['Status'].upper() == "ERROR":
            if "Rows processed(0) is beneath minimum".upper() in row['Details'].upper():
                err_det = "Rows processed(0) is beneath minimum"
            elif "ERROR: cancel by reaper".upper() in row['Details'].upper():
                err_det = "ERROR: cancel by reaper"
            elif "cancelled on user's request".upper() in row['Details'].upper():
                err_det = "Cancelled on user's request"
            elif "ERROR: Insufficient memory to run query".upper() in row['Details'].upper():
                err_det = "ERROR: Insufficient memory to run query"
            elif "ERROR: syntax error at or near".upper() in row['Details'].upper():
                index = row['Details'].upper().find("ERROR: syntax error at or near".upper())            
                err_det = row['Details'][index:]
                index = err_det.upper().find(".".upper())
                err_det = err_det[:index+1]
            elif "consists of duplicate records, rolling back".upper() in row['Details'].upper():
                index = row['Details'].upper().find("consists of duplicate records, rolling back".upper())
                index2 = row['Details'].upper().find("Table".upper())
                if index2<index:
                    err_det = row['Details'][index2:index+len("consists of duplicate records, rolling back")]
                else:
                    err_det = row['Details'][index-50: index+len("consists of duplicate records, rolling back")]
            elif "Error: permission denied for relation".upper() in row['Details'].upper():
                index = row['Details'].upper().find("Error: permission denied for relation".upper())
                err_det = row['Details'][index:]
                index = err_det.upper().find(".".upper())
                err_det = err_det[:index+1]
            else:
                err_det = row['Details'][:311]
            if err_det == "":
                return "Error",row['Details'][:311], 1
            else:
                return "Error",err_det, 1

        else:
            return row['Status'], row['Details'],2

    df1_app['Details'] = df1_app['Details'].astype(str)    
    df1_app['Status'], df1_app['Details'], df1_app['sp'] = zip(*df1_app.apply(msk,axis=1))
    df1_app.rename(columns={'Name':'Table Name'},inplace=True)
    df2_app_f = pd.merge(df1_app,prt,how='left',left_on='Table Name',right_on='tablename').reset_index()
    df1_success.rename(columns={'Name':'Table Name'},inplace=True)
    df1_success = pd.merge(df1_success,prt,how='left',left_on='Table Name',right_on='tablename').reset_index()


    df_error = df1_app[df1_app['Status'] == "Error"]
    print("Printing Error Table from Source")
    print(df_error)

    df2_app_f_t7 = df2_app_f[df2_app_f['Dataset Date'] < start_date]
    df2_app_f_t1 = df2_app_f[df2_app_f['Dataset Date'] == start_date]

    df2_app_f7 = df2_app_f_t7.sort_values(by=['sp','mp','bp','Table Name','Dataset Date','Schedule Hour (Local)'], ascending=[True,True,True,True,True,True]).reset_index()
    df2_app_f1 = df2_app_f_t1.sort_values(by=['sp','mp','bp','Table Name','Dataset Date','Schedule Hour (Local)'], ascending=[True,True,True,True,True,True]).reset_index()
    df1_success = df1_success.sort_values(by=['mp','bp','Table Name','Dataset Date','Schedule Hour (Local)'], ascending=[True,True,True,True,True]).reset_index()

    df2_app_f7['Sl No.'] = df2_app_f7.index + 1
    df2_app_f1['Sl No.'] = df2_app_f1.index + 1
    df1_success['Sl No.'] = df1_success.index + 1


    #SLA start
    # Read sla file to dataframe
    sla_df = pd.read_excel("sla_april.xlsx")
    print(sla_df)
    # Remove [E], [L] from table names

    # df['derived_column'] = df['original_column'].str.replace(r'\[E\]', '').str.replace(r'\[L\]', '').str.strip()
    df1_success['Extr_Table'] = df1_success['Table Name'].apply(lambda x: x.replace('[E]', '').replace('[L]', '').strip())
    df2_app_f1['Extr_Table'] = df2_app_f1['Table Name'].apply(lambda x: x.replace('[E]', '').replace('[L]', '').strip())
    df2_app_f7['Extr_Table'] = df2_app_f7['Table Name'].apply(lambda x: x.replace('[E]', '').replace('[L]', '').strip())

    # Left join and take SLAs
    df1_success = pd.merge(df1_success,sla_df,how='left',left_on='Extr_Table',right_on='Table')
    df2_app_f1 = pd.merge(df2_app_f1,sla_df,how='left',left_on='Extr_Table',right_on='Table')
    df2_app_f7 = pd.merge(df2_app_f7,sla_df,how='left',left_on='Extr_Table',right_on='Table')
    # fill 0 for null SLAs
    df1_success['SLA'] = df1_success['SLA'].fillna(0)
    df2_app_f1['SLA'] = df2_app_f1['SLA'].fillna(0)
    df2_app_f7['SLA'] = df2_app_f7['SLA'].fillna(0)
    #df1_success['Dataset Date'] = df1_success['Dataset Date'].astype('datetime64[ms]')
    # Manipulate datetime columns
    # Success
    df1_success['Dataset Datetime'] = df1_success['Dataset Date'].astype('datetime64[ms]')
    df1_success['End Date (PST)'] = df1_success['End Date (PST)'].astype('datetime64[ms]')
    #df1_success['Actual End Time (IST)'] = df1_success['End Date (PST)'] + timedelta(hours=12, minutes=30)
    #df1_success['Actual End Time (IST)'] = df1_success['End Date (PST)'].dt.tz_localize(pdt).dt.tz_convert(ist)
    df1_success['Actual End Time (IST)'] = (df1_success['End Date (PST)']
        .dt.tz_localize('America/Los_Angeles')  # This will interpret the time as PDT/PST appropriately
        .dt.tz_convert('Asia/Kolkata')          # Convert to IST
        .dt.tz_localize(None)) 
    consume_current_timestamp = time.time()
    dtm_current = datetime.datetime.fromtimestamp(consume_current_timestamp)
    print("Current Time")
    print(dtm_current)
    dtm_current_ft = dtm_current.strftime("%Y-%m-%d")
    print(dtm_current_ft)
    starttime2 = dt_object1
    print("Time taken for excecution")
    difference = dtm_current - starttime2
    print(difference)
    difference_in_hours = difference.total_seconds() / 3600  # Convert seconds to hours
    print(f"Difference in hours: {difference_in_hours:.2f} hours")


    try:
        df1_success['Expected End Time (IST)'] = df1_success.apply(
            lambda row: row['Dataset Datetime'] + timedelta(hours=24) + timedelta(hours=row['SLA']), 
            axis=1
        )
    except:
        df1_success['Expected End Time (IST)'] = df1_success['Dataset Datetime'] + pd.Timedelta(hours=24) + pd.to_timedelta(df1_success['SLA'], unit='hours')

    try:
        df1_success['ElapsedHours'] = df1_success.apply(
        lambda row: ((row['Actual End Time (IST)'] - (row['Dataset Datetime'] + timedelta(hours=24)) ).total_seconds() / 3600),
        axis=1
        )
    except:
        df1_success['ElapsedHours'] = (df1_success['Actual End Time (IST)'] - (df1_success['Dataset Datetime'] + pd.Timedelta(hours=24))).dt.total_seconds() / 3600


    # T-1
    df2_app_f1['Dataset Datetime'] = df2_app_f1['Dataset Date'].astype('datetime64[ms]')
    # df2_app_f1['End Date (PST)'] = df2_app_f1['End Date (PST)'].astype('datetime64[ms]')
    # df2_app_f1['Actual End Time (IST)'] = df2_app_f1['End Date (PST)'] + timedelta(hours=12, minutes=30)
    try:
        df2_app_f1['Expected End Time (IST)'] = df2_app_f1.apply(
            lambda row: row['Dataset Datetime'] + timedelta(hours=24) + timedelta(hours=row['SLA']), 
            axis=1
        )
    except:
        df2_app_f1['Expected End Time (IST)'] = df2_app_f1['Dataset Datetime'] + pd.Timedelta(hours=24) + pd.to_timedelta(df2_app_f1['SLA'], unit='hours')
    try:
        df2_app_f1['ElapsedHours'] = df2_app_f1.apply(
            lambda row: ((dtm_current - (row['Dataset Datetime'] + timedelta(hours=24)) ).total_seconds() / 3600),
            axis=1
        )
    except:
        df2_app_f1['ElapsedHours'] = (dtm_current - (df2_app_f1['Dataset Datetime'] + pd.Timedelta(hours=24))).dt.total_seconds() / 3600


    # T-2 to T-7
    df2_app_f7['Dataset Datetime'] = df2_app_f7['Dataset Date'].astype('datetime64[ms]')
    try:
        df2_app_f7['End Date (PST)'] = datetime.datetime.now()
        print(df2_app_f7['End Date (PST)'])
        df2_app_f7['Actual End Time (IST)'] = df2_app_f7['End Date (PST)'] + timedelta(hours=12, minutes=30)
        df2_app_f7['ElapsedHours'] = df2_app_f7.apply(
            lambda row: ((dtm_current - (row['Dataset Datetime'] + timedelta(hours=24)) ).total_seconds() / 3600),
            axis=1
        )
    except:
        df2_app_f7['ElapsedHours'] = (dtm_current - (df2_app_f7['Dataset Datetime'] + pd.Timedelta(hours=24))).dt.total_seconds() / 3600



    try:
        df2_app_f7['Expected End Time (IST)'] = df2_app_f7.apply(
            lambda row: row['Dataset Datetime'] + timedelta(hours=24) + timedelta(hours=row['SLA']), 
            axis=1
        )
    except:
        df2_app_f7['Expected End Time (IST)'] = df2_app_f7['Dataset Datetime'] + pd.Timedelta(hours=24) + pd.to_timedelta(df2_app_f7['SLA'], unit='hours')


    # Flag SLA breaching
    # Success
    df1_success['SLA_FLAG'] = df1_success.apply(lambda row: 'Not Breached' if row['Actual End Time (IST)'] <= row['Expected End Time (IST)'] else 'Breached', axis=1)
    df1_success['SLA'] = df1_success['SLA'].astype(str).str.replace('.0','')
    df1_success['Expected End Time (IST)'] = pd.to_datetime(df1_success['Expected End Time (IST)']).dt.strftime('%Y-%m-%d %H:%M:%S')
    df1_success.to_excel("df1_success_sla_flag.xlsx")
    print(df1_success[['Table Name','Dataset Date','SLA','End Date (PST)','Actual End Time (IST)','Expected End Time (IST)','SLA_FLAG']])
    print(df1_success.dtypes)

    #T-1
    df2_app_f1['SLA'] = df2_app_f1['SLA'].astype(str).str.replace('.0','')
    df2_app_f1['SLA_FLAG'] = df2_app_f1.apply(lambda row: 'Not Yet Breached' if datetime.datetime.now() <= row['Expected End Time (IST)'] else 'Breached', axis=1)
    df2_app_f1['Expected End Time (IST)'] = pd.to_datetime(df2_app_f1['Expected End Time (IST)']).dt.strftime('%Y-%m-%d %H:%M:%S')
    df2_app_f1.to_excel("df2_app_f1_sla_flag.xlsx")
    print(df2_app_f1[['Table Name','Dataset Date','SLA','Expected End Time (IST)','SLA_FLAG']])
    print(df2_app_f1.dtypes)

    # T-2 to T-7
    df2_app_f7['SLA'] = df2_app_f7['SLA'].astype(str).str.replace('.0','')
    try:
        df2_app_f7['SLA_FLAG'] = df2_app_f7.apply(lambda row: 'Not Yet Breached' if datetime.datetime.now() <= row['Expected End Time (IST)'] else 'Breached', axis=1)
    except:
        df2_app_f7['SLA_FLAG'] = ""

    df2_app_f7['Expected End Time (IST)'] = pd.to_datetime(df2_app_f7['Expected End Time (IST)']).dt.strftime('%Y-%m-%d %H:%M:%S')
    df2_app_f7.to_excel("df2_app_f7_sla_flag.xlsx")
    print(df2_app_f7[['Table Name','Dataset Date','SLA','Expected End Time (IST)','SLA_FLAG']])
    print(df2_app_f7.dtypes)



    print("Success")
    print(df1_success[['Sl No.','Table Name','Status','Dataset Date','SLA','Expected End Time (IST)','Actual End Time (IST)','SLA_FLAG']].head(20))
    print("T-1")
    print(df2_app_f1[['Sl No.','Table Name','Status','Dataset Date','SLA','Expected End Time (IST)','SLA_FLAG']].head(20))
    print(">=T-2")
    print(df2_app_f7[['Sl No.','Table Name','Status','Dataset Date','SLA','Expected End Time (IST)','SLA_FLAG']].head(20))
    #SLA end


    df2_app_f7 = df2_app_f7[['Sl No.','Table Name','Status','Dataset Date','Schedule Hour (Local)','JobRunLink','JobHistoryLink','Sat_Dependency','Details','SLA','ElapsedHours','Expected End Time (IST)','SLA_FLAG']]
    df2_app_f1 = df2_app_f1[['Sl No.','Table Name','Status','Dataset Date','Schedule Hour (Local)','JobRunLink','JobHistoryLink','Sat_Dependency','Details','SLA','ElapsedHours','Expected End Time (IST)','SLA_FLAG']]
    df1_success = df1_success[['Sl No.','Table Name','Status','Dataset Date','Schedule Hour (Local)','JobRunLink','JobHistoryLink','SLA','ElapsedHours','Expected End Time (IST)','Actual End Time (IST)','SLA_FLAG']]

    df2_app_ft12 = pd.concat([df2_app_f1, df2_app_f7], ignore_index=True)

    df1_success_t12 = df1_success
    df1_success_t12['Sat_Dependency'] = None
    df1_success_t12['Details'] = None
    df2_app_ft12['Actual End Time (IST)'] = None
    df2_app_ft12 = df2_app_ft12[['Sl No.','Table Name','Status','Dataset Date','Schedule Hour (Local)','JobRunLink','JobHistoryLink','Sat_Dependency','Details','SLA','ElapsedHours','Expected End Time (IST)','Actual End Time (IST)','SLA_FLAG']]
    df1_success_t12 = df1_success_t12[['Sl No.','Table Name','Status','Dataset Date','Schedule Hour (Local)','JobRunLink','JobHistoryLink','Sat_Dependency','Details','SLA','ElapsedHours','Expected End Time (IST)','Actual End Time (IST)','SLA_FLAG']]
    df2_app_t12 = pd.concat([df1_success_t12, df2_app_ft12], ignore_index=True)


    df2_app_t12['Extr_Table'] = df2_app_t12['Table Name'].apply(lambda x: x.replace('[E]', '').replace('[L]', '').strip())

    # Left join and take SLAs
    df2_app_t12 = pd.merge(df2_app_t12,df_t12,how='inner',left_on='Extr_Table',right_on='Table').drop_duplicates().reset_index()

    df2_app_t12['Sl No.'] = df2_app_t12.index + 1
    df2_app_t12 = df2_app_t12[['Sl No.','Table Name','Status','Dataset Date','Schedule Hour (Local)','JobRunLink','JobHistoryLink','Sat_Dependency','Details','SLA','ElapsedHours','Expected End Time (IST)','Actual End Time (IST)','SLA_FLAG']]

    df2_app_f7_xl = df2_app_f7
    df2_app_f1_xl = df2_app_f1

    df2_app_f1_error = df2_app_f1[df2_app_f1['Status'] == "Error"]
    df2_app_f7_error = df2_app_f7[df2_app_f7['Status'] == "Error"]
    force_app = pd.DataFrame()
    force_exc = pd.DataFrame()
    force_revisited = pd.DataFrame()
    loop1_exc2err  = pd.DataFrame()
    force_app2 = pd.DataFrame()
    force_app2_xl = pd.DataFrame()
    df_force_success = pd.DataFrame()

    table_force_success = df2_app_f7[df2_app_f7['Table Name'].str.startswith(tuple(forcesuccess_list))]
    if not table_force_success.empty:
        table_force_success = table_force_success.sort_values(by=['Table Name','Dataset Date'], ascending=[True,False]).reset_index()
        table_force_success['Sl No.'] = table_force_success.index + 1
        force_unique = table_force_success[['Table Name','JobHistoryLink']].drop_duplicates().reset_index()
        print(force_unique)
        i=0
        force_app = pd.DataFrame()
        force_exc = pd.DataFrame()
        force_revisited = pd.DataFrame()
        loop1_exc2err  = pd.DataFrame()
        force_app2 = pd.DataFrame()
        force_app2_xl = pd.DataFrame()
        df_force_success = pd.DataFrame()
        #Enter FOR loop1
        for index, row in force_unique.iterrows():
            name = row['Table Name']
            link_h = row['JobHistoryLink']
            i+=1
            print(i)
            response = make_request(link_h)
            soup = BeautifulSoup(response.content, 'html.parser')
            # Find the table you're interested in
            table = soup.find('table', class_='tablesorter')
            # Extract the table data using pandas
            table_data = pd.read_html(str(table))
            # Display the table data
            table_rows = table_data[0]

            # Fetching T-1 success jobruns on ideaprod_pmnt2 (No further operations except appending and sending as an attachment on mail)
            table_rows = table_rows[(table_rows['Dataset Date']>=end_date) & (table_rows['Dataset Date']<=start_date)]
            print(table_rows)
            #table_success = table_success[(table_success['Database'] == "ideaprod_pmnt2") | table_success['Database'].isnull() | table_success['Database'].isna()]
            table_rows['Table Name'] = name
            table_rows['JobHistoryLink'] = link_h
            table_rows['JobRunLink'] = "https://datacentral.a2z.com/console?action=jobrun_details&jobrun_id=" + table_rows['Job Run'].astype(str)
            table_rows['Sl No'] = i
            tablerow = table_rows[['Sl No','Job Run','Table Name','Dataset Date','Status','Schedule Hour (Local)','JobRunLink','JobHistoryLink','End Date (PST)']]
            print(table_rows)
            tablelatestsuccess = tablerow[(tablerow['Status'] == 'Success')]
            tablelatestsuccess = tablelatestsuccess.sort_values(['Table Name', 'Dataset Date'], ascending=[True, False])
            print(tablelatestsuccess)
            tablelatestsuccess = tablelatestsuccess[['Table Name', 'Dataset Date']].drop_duplicates().head(1)
            print(tablelatestsuccess)
            tablerow2 = tablerow.merge(tablelatestsuccess[['Table Name', 'Dataset Date']], on='Table Name', suffixes=('', '_latest'))
            print(tablerow2)
            print(tablerow2.dtypes)
            tablerow2 = tablerow2[tablerow2['Dataset Date'] < tablerow2['Dataset Date_latest']]
            print(tablerow2)
            tablerownotsuccess = tablerow2[(tablerow2['Status'] != 'Success')]
            print(tablerownotsuccess)
            df_tablerownotsuccess = pd.concat([tablerownotsuccess],ignore_index=True)
            force_app = force_app._append(df_tablerownotsuccess,ignore_index=True)
            print(force_app)
            force_app2 = force_app[['Job Run','Table Name','Dataset Date','Status','Schedule Hour (Local)','JobRunLink','JobHistoryLink','End Date (PST)']]
            force_app2 = force_app2.reset_index()
            force_app2['Sl No.'] = force_app2.index + 1
            force_app2 = force_app2[['Sl No.','Job Run','Table Name','Dataset Date','Status','Schedule Hour (Local)','JobRunLink','JobHistoryLink','End Date (PST)']]
            force_app2 = force_app2.sort_values(by=['Sl No.'], ascending=[True])
            force_app2_xl = force_app2
            force_app2_xl.to_excel("All_force_success_jobs.xlsx")
            force_app2['JobRunLink'] = force_app2['JobRunLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
            force_app2['JobHistoryLink'] = force_app2['JobHistoryLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
            def highlightserr(rows_er):
                if rows_er['Status']=='Error':
                    return ['background-color: red; color: white;'] * len(rows_er)
                elif rows_er['Status']=='Waiting for Dependencies' or rows_er['Status']=='Waiting for Resources':
                    return ['background-color: linen; color: black;'] * len(rows_er) #lavender
                elif rows_er['Status']=='Executing':
                    return ['background-color: orange; color: black;'] * len(rows_er)
                else:
                    return [''] * len(rows_er)
            errrr = force_app2.style.apply(highlightserr, axis=1)
            table_style2 = [
            {'selector': 'table',
            'props': [
                ('border-collapse', 'collapse')
                #('border-radius', '500px')  # Adjust the radius as needed
            ]},
            {'selector': 'th, td',
            'props': [
                ('border', '1px solid black'),
                ('padding', '8px')
                #('border-radius', '500px')
            ]},
            {'selector': 'th',
            'props': [
                ('background-color', 'purple') #bisque
            ]}]
            errrr = errrr.set_table_styles(table_style2)
            errrr = errrr.set_table_attributes('style="border: 1px solid black;"')
            errrr.hide()
            bodyerr = errrr.to_html()
            print("Printing force success wala job")
            print(force_app2_xl)
            if not force_app2_xl.empty:      
                i=0  
                for index, row in force_app2_xl.iterrows():
                    try:
                        from webbrowser import get
                        from selenium import webdriver
                        from selenium.webdriver.chrome.service import Service
                        from webdriver_manager.chrome import ChromeDriverManager # type: ignore
                        from selenium.webdriver.chrome.options import Options
                        from selenium.webdriver.common.by import By
                        from time import sleep
                        import time
                        #Chrome Setup Start

                        chrome_options = Options()
                        chrome_options.add_argument("--headless")
                        chrome_options.add_argument("--window-size=1920x1080")
                        chrome_options.add_argument("--disable-notifications")
                        chrome_options.add_experimental_option("prefs", {
                            "download.default_directory": "<path_to_download_default_directory>",
                            "download.prompt_for_download": False,
                            "download.directory_upgrade": True, 
                            "safebrowsing_for_trusted_sources_enabled": False,
                            "safebrowsing.enabled": False
                        })

                        #fetching midway cookie
                        def get_mwinit_cookie():
                            MidwayConfigDir = os.path.join(os.path.expanduser("~"), ".midway")
                            MidwayCookieJarFile = os.path.join(MidwayConfigDir, "cookie")
                            fields = []
                            keyfile = open(MidwayCookieJarFile, "r")
                            for line in keyfile:
                                # parse the record into fields (separated by whitespace)
                                fields = line.split()
                                if len(fields) != 0:
                                    # get the yubi session token and expire time
                                    if fields[0] == "#HttpOnly_midway-auth.amazon.com":
                                        session_token = fields[6].replace("\n", "")
                                        expires = fields[4]
                                    # get the user who generated the session token
                                    elif fields[0] == "midway-auth.amazon.com":
                                        username = fields[6].replace("\n", "")
                            keyfile.close()
                            # make sure the session token hasn't expired
                            if time.gmtime() > time.gmtime(int(expires)):
                                raise SystemError("Your Midway token has expired. Run mwinit to renew")
                            # construct the cookie value required by calls to k2
                            cookie = {"username": username, "session": session_token}
                            return cookie

                        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
                        #Adding cookie to driver
                        midway_url = 'https://midway-auth.amazon.com'
                        cookie = get_mwinit_cookie()
                        driver.get(midway_url)
                        cookie_dict1 = {'domain': '.midway-auth.amazon.com',
                                        'name': 'user_name',
                                        'value': cookie['username'],
                                        'path': '/',
                                        'httpOnly': False,
                                        'secure': True}

                        cookie_dict2 = {
                            'domain': '.midway-auth.amazon.com',
                            'name': 'session',
                            'value': cookie['session'],
                            'path': '/',
                            'httpOnly': True,
                            'secure': True}

                        driver.add_cookie(cookie_dict1)
                        driver.add_cookie(cookie_dict2)

                        match = False
                        while not match:
                            driver.get(midway_url)
                            if driver.current_url == 'https://midway-auth.amazon.com/':
                                match = True
                            sleep(1)
                            driver.refresh()
                        driver.maximize_window()
                        driver.implicitly_wait(60)
                        linkt="https://datacentral.a2z.com/dw-platform/servlet/dwp/template/IdRedirector.vm/id/J21108063"
                        print(linkt)
                        driver.get(str(linkt))
                        driver.implicitly_wait(60)
                        get_title = driver.title
                        print(get_title)
                        driver.save_screenshot("dcsbv.png")
                        #Chrome Setup End
                        #Chrome Setup End
                        explorer = 'https://datanet-service.amazon.com/explorer/index.html#'
                        driver.get(explorer)
                        sleep(2)
                        driver.save_screenshot("dcsbv.png")
                        #driver.find_element(By.XPATH,"/html/body/div[2]/ul/li[88]/span/a").click()
                        driver.find_element(By.ID,"forceJobRunStatus_link").click()
                        sleep(5)
                        driver.save_screenshot("dcsbv.png")
                
                        table_rowserr = pd.DataFrame()
                        name2 = row['Table Name']
                        stat = row['Status']
                        dsdate = row['Dataset Date']
                        schedule = row['Schedule Hour (Local)']
                        link_jr = row['JobRunLink']
                        link_jh = row['JobHistoryLink']
                        res_et = row['End Date (PST)']
                        driver.implicitly_wait(10)
                        # query_string = link_jr.split('?')[1].split('&')
                        jobrun_id = row['Job Run']

                        # for param in query_string:
                        #     if param.startswith('jobrun_id='):
                        #         jobrun_id = param.split('=')[1]
                        #         break

                        print(jobrun_id)
                        input_json = '{ "jobRunId": ' + str(jobrun_id)+ ' , "status": "SUCCESS", "reason": "Succeeding Job Run Is Complete"}'
                        print("Input Json")
                        print(input_json)
                        driver.find_element(By.XPATH,f"{fs_text}").clear()
                        sleep(3)
                        driver.save_screenshot("dcsbv.png")
                        driver.find_element(By.XPATH,f"{fs_text}").send_keys(input_json)
                        sleep(3)
                        driver.save_screenshot("dcsbv.png")
                        driver.find_element(By.XPATH,f"{fs_submit}").click()
                        sleep(3)
                        driver.save_screenshot("dcsbv.png")
                        # our = driver.find_element(By.CLASS_NAME,"spacer")
                        output_json = driver.find_element(By.XPATH,f"{fs_response}")
                        # print(output_json)
                        # print(our)
                        import json
                        print(output_json.text)
                        data = json.loads(output_json.text)

                        print(data['jobRunId'])
                        link_njrid = data['jobRunId']
                        res_det = "Job run has been forced success"
                        
                        # print(details)
                        table_rowserr['Sl No.'] = pd.Series([i] * 1)
                        table_rowserr['Table Name'] = pd.Series([name2] * 1)
                        table_rowserr['JobHistoryLink'] = pd.Series([link_jh] * 1)
                        table_rowserr['JobRunLink'] = pd.Series([link_jr] * 1)
                        table_rowserr['Dataset Date'] = pd.Series([dsdate] * 1)
                        table_rowserr['Status'] = pd.Series([stat] * 1)
                        table_rowserr['Details'] = pd.Series([res_det] * 1)
                        table_rowserr['Schedule Hour (Local)'] = pd.Series([schedule] * 1)
                        table_rowserr['End Date (PST)'] = pd.Series([res_et] * 1)
                        table_rowserr['Job Run'] = pd.Series([jobrun_id] * 1)

                        table_rowserr = table_rowserr[['Sl No.','Job Run','Table Name','Dataset Date','Status','Details','Schedule Hour (Local)','JobRunLink','JobHistoryLink','End Date (PST)']]

                        print(table_rowserr)
                        print(table_rowserr.dtypes)
                        loop_success_force = pd.concat([table_rowserr],ignore_index=True)
                        df_force_success = df_force_success._append(loop_success_force,ignore_index=True)
                        print(df_force_success)
                    except Exception as er:
                        excerr = pd.DataFrame()
                        name2 = row['Table Name']
                        stat = row['Status']
                        dsdate = row['Dataset Date']
                        schedule = row['Schedule Hour (Local)']
                        link_jr = row['JobRunLink']
                        link_jh = row['JobHistoryLink']
                        res_det = "Job run could not be forced success"

                        excerr['Sl No'] = pd.Series([i] * 1)
                        excerr['Table Name'] = pd.Series([name2] * 1)
                        excerr['JobHistoryLink'] = pd.Series([link_jh] * 1)
                        excerr['JobRunLink'] = pd.Series([link_jr] * 1)
                        excerr['Dataset Date'] = pd.Series([dsdate] * 1)
                        excerr['Status'] = pd.Series([stat] * 1)
                        excerr['Details'] = pd.Series([res_det] * 1)
                        excerr['Schedule Hour (Local)'] = pd.Series([schedule] * 1)
                        excerr['Details'] = pd.Series([res_det] * 1)
                        excerr['Exception_Details'] =  pd.Series([str(er)] * 1) 
                        print(excerr)
                        loop1_exc2_ferr = pd.concat([excerr],ignore_index=True)
                        loop1_exc2err = loop1_exc2err._append(loop1_exc2_ferr,ignore_index=True)
                
                loop1_exc2err.to_excel("exception_in_force_success.xlsx")
                df_force_success.to_excel("forcedsuccessjobs.xlsx") 
                if not loop1_exc2err.empty:
                    loop1_exc2err['JobRunLink'] = loop1_exc2err['JobRunLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
                    loop1_exc2err['JobHistoryLink'] = loop1_exc2err['JobHistoryLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
                print("Printing restart after loop")
                print(df_force_success)
                print("Printing exception in restart after loop")
                print(loop1_exc2err)
                if not df_force_success.empty:
                    df_force_success['JobRunLink'] = df_force_success['JobRunLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
                    df_force_success['JobHistoryLink'] = df_force_success['JobHistoryLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
                def highlightserr2(rows_er2):
                    if rows_er2['Status']=='Error':
                        return ['background-color: red; color: white;'] * len(rows_er2)
                    elif rows_er2['Status']=='Waiting for Dependencies' or rows_er2['Status']=='Waiting for Resources':
                        return ['background-color: linen; color: black;'] * len(rows_er2) #lavender
                    elif rows_er2['Status']=='Executing':
                        return ['background-color: orange; color: black;'] * len(rows_er2)
                    else:
                        return [''] * len(rows_er2)
                res = df_force_success.style.apply(highlightserr2, axis=1)
                table_style3 = [
                {'selector': 'table',
                'props': [
                    ('border-collapse', 'collapse')
                    #('border-radius', '500px')  # Adjust the radius as needed
                ]},
                {'selector': 'th, td',
                'props': [
                    ('border', '1px solid black'),
                    ('padding', '8px')
                    #('border-radius', '500px')
                ]},
                {'selector': 'th',
                'props': [
                    ('background-color', 'purple') #bisque
                ]}]
                res = res.set_table_styles(table_style3)
                res = res.set_table_attributes('style="border: 1px solid black;"')
                res.hide()
                bodyerrres = res.to_html()

                def highlightserr2x(rows_er2x):
                    if rows_er2x['Status']=='Error':
                        return ['background-color: red; color: white;'] * len(rows_er2x)
                    elif rows_er2x['Status']=='Waiting for Dependencies' or rows_er2x['Status']=='Waiting for Resources':
                        return ['background-color: linen; color: black;'] * len(rows_er2x) #lavender
                    elif rows_er2x['Status']=='Executing':
                        return ['background-color: orange; color: black;'] * len(rows_er2x)
                    else:
                        return [''] * len(rows_er2x)
                erx = loop1_exc2err.style.apply(highlightserr2x, axis=1)
                table_style4 = [
                {'selector': 'table',
                'props': [
                    ('border-collapse', 'collapse')
                    #('border-radius', '500px')  # Adjust the radius as needed
                ]},
                {'selector': 'th, td',
                'props': [
                    ('border', '1px solid black'),
                    ('padding', '8px')
                    #('border-radius', '500px')
                ]},
                {'selector': 'th',
                'props': [
                    ('background-color', 'purple') #bisque
                ]}]
                erx = erx.set_table_styles(table_style4)
                erx = erx.set_table_attributes('style="border: 1px solid black;"')
                erx.hide()
                bodyerrx = erx.to_html()

                #Sending mail for error and restart
                outlook = win32.Dispatch('outlook.application')
                mail = outlook.CreateItem(0)
                #Outlook setup starts
                if teamss != '111':
                    if teamss == '1':
                        # mail.To = 'kumraov'+'@amazon.com'
                        # mail.To = 'mdsagkha'+'@amazon.com;' + 'kumraov'+'@amazon.com;'
                        mail.To = 'inpay-dp-team@amazon.com;'  + 'mdsagkha'+'@amazon.com;'
                        mail.cc='sastry' + '@amazon.com;' +  'plammojo'+'@amazon.com;iankranj@amazon.com'
                    
                    else:
                        mail.cc = login + '@amazon.com;'

                    mail.Body = 'Hi team,'+'\n\n'+'Please find the updated sheet for forced success P0 tables (' + run + ' run).'+'\n'
                    # mail.HTMLBody += "<b><font size='1'>Notes for SLA_FLAG:</font></b>"+'\n\n'+"<br><i><font size='1'>Breached: Breached SLA with any jobrun status (success, error, wfd & wfr)</font></i>"
                    # mail.HTMLBody += "<br><i><font size='1'>Not Breached: Job run status is success and did not breach SLA</font></i>"+'\n'
                    # mail.HTMLBody += "<br><i><font size='1'>Not Yet Breached: Job run status is not success and did not breach SLA</font></i><br>"+'\n'

                    # mail.HTMLBody += "There were some code changes done, hence sending updated mail<br><br>"
                    # mail.HTMLBody += "Sheet1 contains details from t-7 to t-4 and Sheet2 from t-3 to t-1<br><br>"
                    mail.HTMLBody += "<b><i><u><font size='5'>All Force Success Required Tables:</font></u></i></b><br>"
                    # mail.HTMLBody += "<br><br>"
                    mail.HTMLBody += bodyerr
                    # mail.HTMLBody += "<br><br>"
                    mail.HTMLBody += "<br><b><i><u><font size='5'>Forced Success Tables:</font></u></i></b><br>"
                    # mail.HTMLBody += "<br><br>"
                    mail.HTMLBody += bodyerrres
                    mail.HTMLBody += "<br><b><i><u><font size='5'>Exception details during forcing success:</font></u></i></b><br>"
                    # mail.HTMLBody += "<br><br>"
                    mail.HTMLBody += bodyerrx
                    mail.HTMLBody += f"<b><br>Thanks,<br>{mailname}</b>"
                    #mail.Subject = 'Data Monitoring Sheet - Platform Team'
                    mail.Subject = 'Force Success!(' + run + ' run): Data Monitoring Sheet for P0 Tables - Platform Team'
                    login = os.getlogin()
                    attachment1  = "C:\\Users\\"+login+"\\Downloads\\dn_proj_reg\\"+"All_force_success_jobs.xlsx"
                    attachment2  = "C:\\Users\\"+login+"\\Downloads\\dn_proj_reg\\"+"forcedsuccessjobs.xlsx"
                    attachment3  = "C:\\Users\\"+login+"\\Downloads\\dn_proj_reg\\"+"exception_in_force_success.xlsx"
                    mail.Attachments.Add(attachment1)
                    mail.Attachments.Add(attachment2)
                    mail.Attachments.Add(attachment3)
                    mail.Send()
                    print('mail gaya')
                    #Outlook setup ends
                else:
                            #Sending mail for error and restart
                    outlook = win32.Dispatch('outlook.application')
                    mail = outlook.CreateItem(0)
                    #Outlook setup starts
                    if teamss != '111':
                        if teamss == '1':
                        # mail.To = 'kumraov'+'@amazon.com'
                        # mail.To = 'mdsagkha'+'@amazon.com;' + 'kumraov'+'@amazon.com;'
                            mail.To = 'inpay-dp-team@amazon.com;'  + 'mdsagkha'+'@amazon.com;'
                            mail.cc='sastry' + '@amazon.com;' + 'plammojo'+'@amazon.com;iankranj@amazon.com'
                        
                        else:
                            mail.cc = login + '@amazon.com;'

                        mail.Body = 'Hi team,'+'\n\n'+'Please find the updated sheet for forced success P0 tables (' + run + ' run).'+'\n'
                        # mail.HTMLBody += "<b><font size='1'>Notes for SLA_FLAG:</font></b>"+'\n\n'+"<br><i><font size='1'>Breached: Breached SLA with any jobrun status (success, error, wfd & wfr)</font></i>"
                        # mail.HTMLBody += "<br><i><font size='1'>Not Breached: Job run status is success and did not breach SLA</font></i>"+'\n'
                        # mail.HTMLBody += "<br><i><font size='1'>Not Yet Breached: Job run status is not success and did not breach SLA</font></i><br>"+'\n'

                        # mail.HTMLBody += "There were some code changes done, hence sending updated mail<br><br>"
                        # mail.HTMLBody += "Sheet1 contains details from t-7 to t-4 and Sheet2 from t-3 to t-1<br><br>"
                        mail.HTMLBody += "<b><i><u><font size='5'>All Errored Out Tables:</font></u></i></b><br>"
                        # mail.HTMLBody += "<br><br>"
                        mail.HTMLBody += bodyerr
                        
                        mail.HTMLBody += f"<b><br>Thanks,<br>{mailname}</b>"
                        #mail.Subject = 'Data Monitoring Sheet - Platform Team'
                        mail.Subject = 'Force Success!(' + run + ' run): Data Monitoring Sheet for P0 Tables - Platform Team'
                        login = os.getlogin()
                        attachment1  = "C:\\Users\\"+login+"\\Downloads\\dn_proj_reg\\"+"All_force_success_jobs.xlsx"
                        mail.Attachments.Add(attachment1)
                        mail.Send()
                        print('mail gaya')
                        #Outlook setup ends
    print(force_app)

    print(table_force_success)

    df_error2 = pd.concat([df2_app_f1_error, df2_app_f7_error], ignore_index=True)
    print("Working till here")
    print(df_error2)
    df_restarted = pd.DataFrame()
    loop_restarted = pd.DataFrame()
    loop1_exc2_ferr = pd.DataFrame()
    loop1_exc2err = pd.DataFrame()
    if not df_error2.empty:

        df_error2 = df_error2[['Table Name','Status','Dataset Date','Schedule Hour (Local)','JobRunLink','JobHistoryLink','Sat_Dependency','Details','SLA','ElapsedHours','Expected End Time (IST)','SLA_FLAG']]
        df_error2 = df_error2.reset_index()
        df_error2['Sl No.'] = df_error2.index + 1
        df_error2 = df_error2[['Sl No.','Table Name','Status','Dataset Date','Schedule Hour (Local)','JobRunLink','JobHistoryLink','Sat_Dependency','Details','SLA','ElapsedHours','Expected End Time (IST)','SLA_FLAG']]
        df_error2 = df_error2.sort_values(by=['Sl No.'], ascending=[True])
        df_error2_xl = df_error2
        df_error2_xl.to_excel("All_error_jobs.xlsx")
        df_error2['JobRunLink'] = df_error2['JobRunLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
        df_error2['JobHistoryLink'] = df_error2['JobHistoryLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
        def highlightserr(rows_er):
            if rows_er['Status']=='Error':
                return ['background-color: red; color: white;'] * len(rows_er)
            elif rows_er['Status']=='Waiting for Dependencies' or rows_er['Status']=='Waiting for Resources':
                return ['background-color: linen; color: black;'] * len(rows_er) #lavender
            elif rows_er['Status']=='Executing':
                return ['background-color: orange; color: black;'] * len(rows_er)
            else:
                return [''] * len(rows_er)
        errrr = df_error2.style.apply(highlightserr, axis=1)
        table_style2 = [
        {'selector': 'table',
        'props': [
            ('border-collapse', 'collapse')
            #('border-radius', '500px')  # Adjust the radius as needed
        ]},
        {'selector': 'th, td',
        'props': [
            ('border', '1px solid black'),
            ('padding', '8px')
            #('border-radius', '500px')
        ]},
        {'selector': 'th',
        'props': [
            ('background-color', 'purple') #bisque
        ]}]
        errrr = errrr.set_table_styles(table_style2)
        errrr = errrr.set_table_attributes('style="border: 1px solid black;"')
        errrr.hide()
        bodyerr = errrr.to_html()
        #df_error2_xl_reaper = df_error2_xl[df_error2_xl['Details'] == "ERROR: cancel by reaper"]
        #df_error2_xl_reaper = df_error2_xl[(df_error2_xl['Details'] == "ERROR: cancel by reaper") | (df_error2_xl['Details'] == "ERROR: Insufficient memory to run query")]
        df_error2_xl_reaper = df_error2_xl[df_error2_xl['Details'].isin([
        "ERROR: cancel by reaper", 
        "ERROR: Insufficient memory to run query"
    ])]

        print("Printing reaper wala job")
        print(df_error2_xl_reaper)
        if not df_error2_xl_reaper.empty:

            i=0
            
            for index, row in df_error2_xl_reaper.iterrows():
                try:
                    from webbrowser import get
                    from selenium import webdriver
                    from selenium.webdriver.chrome.service import Service
                    from webdriver_manager.chrome import ChromeDriverManager # type: ignore
                    from selenium.webdriver.chrome.options import Options
                    from selenium.webdriver.common.by import By
                    from time import sleep
                    import time
                    #Chrome Setup Start

                    chrome_options = Options()
                    chrome_options.add_argument("--headless")
                    chrome_options.add_argument("--window-size=1920x1080")
                    chrome_options.add_argument("--disable-notifications")
                    chrome_options.add_experimental_option("prefs", {
                        "download.default_directory": "<path_to_download_default_directory>",
                        "download.prompt_for_download": False,
                        "download.directory_upgrade": True, 
                        "safebrowsing_for_trusted_sources_enabled": False,
                        "safebrowsing.enabled": False
                    })

                    #fetching midway cookie
                    def get_mwinit_cookie():
                        MidwayConfigDir = os.path.join(os.path.expanduser("~"), ".midway")
                        MidwayCookieJarFile = os.path.join(MidwayConfigDir, "cookie")
                        fields = []
                        keyfile = open(MidwayCookieJarFile, "r")
                        for line in keyfile:
                            # parse the record into fields (separated by whitespace)
                            fields = line.split()
                            if len(fields) != 0:
                                # get the yubi session token and expire time
                                if fields[0] == "#HttpOnly_midway-auth.amazon.com":
                                    session_token = fields[6].replace("\n", "")
                                    expires = fields[4]
                                # get the user who generated the session token
                                elif fields[0] == "midway-auth.amazon.com":
                                    username = fields[6].replace("\n", "")
                        keyfile.close()
                        # make sure the session token hasn't expired
                        if time.gmtime() > time.gmtime(int(expires)):
                            raise SystemError("Your Midway token has expired. Run mwinit to renew")
                        # construct the cookie value required by calls to k2
                        cookie = {"username": username, "session": session_token}
                        return cookie

                    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
                    #Adding cookie to driver
                    midway_url = 'https://midway-auth.amazon.com'
                    cookie = get_mwinit_cookie()
                    driver.get(midway_url)
                    cookie_dict1 = {'domain': '.midway-auth.amazon.com',
                                    'name': 'user_name',
                                    'value': cookie['username'],
                                    'path': '/',
                                    'httpOnly': False,
                                    'secure': True}

                    cookie_dict2 = {
                        'domain': '.midway-auth.amazon.com',
                        'name': 'session',
                        'value': cookie['session'],
                        'path': '/',
                        'httpOnly': True,
                        'secure': True}

                    driver.add_cookie(cookie_dict1)
                    driver.add_cookie(cookie_dict2)

                    match = False
                    while not match:
                        driver.get(midway_url)
                        if driver.current_url == 'https://midway-auth.amazon.com/':
                            match = True
                        sleep(1)
                        driver.refresh()
                    driver.maximize_window()
                    driver.implicitly_wait(60)
                    linkt="https://datacentral.a2z.com/dw-platform/servlet/dwp/template/IdRedirector.vm/id/J21108063"
                    print(linkt)
                    driver.get(str(linkt))
                    driver.implicitly_wait(60)
                    get_title = driver.title
                    print(get_title)
                    driver.save_screenshot("dcsbv.png")
                    #Chrome Setup End
                    explorer = 'https://datanet-service.amazon.com/explorer/index.html#'
                    driver.get(explorer)
                    sleep(2)
                    driver.save_screenshot("dcsbv.png")
                    #driver.find_element(By.XPATH,"/html/body/div[2]/ul/li[88]/span/a").click()
                    driver.find_element(By.ID,"restartJobRun_link").click()
                    sleep(2)
                    driver.save_screenshot("dcsbv.png")
            
                    table_rowserr = pd.DataFrame()
                    name2 = row['Table Name']
                    stat = row['Status']
                    dsdate = row['Dataset Date']
                    schedule = row['Schedule Hour (Local)']
                    link_jr = row['JobRunLink']
                    link_jh = row['JobHistoryLink']
                    res_sla = row['SLA']
                    res_sla_elap = row['ElapsedHours']
                    res_et = row['Expected End Time (IST)']
                    res_fl = row['SLA_FLAG']
                    detail_text = row['Details']
                    driver.implicitly_wait(10)
                    #query_string = link_jr.split('?')[1].split('&')
                    # query_string = link_jr.split('/')[-1]
                    #jobrun_id = None

                    # for param in query_string:
                    #     if param.startswith('jobrun_id='):
                    #         jobrun_id = param.split('=')[1]
                    #         break
                    print("Print link_jr")
                    print(link_jr)
                    start_pos = link_jr.find("https://")
                    end_pos = link_jr.find('"', start_pos + 1)
                    url = link_jr[start_pos:end_pos]

                    jobrun_id = url.split('/')[-1]

                    print(jobrun_id)
                    if detail_text == "ERROR: cancel by reaper":
                        input_json = '{ "jobRunId": ' + str(jobrun_id)+ ' , "reason": "ERROR: cancel by reaper"}'
                    if detail_text == "ERROR: Insufficient memory to run query":
                        input_json = '{ "jobRunId": ' + str(jobrun_id)+ ' , "reason": "ERROR: Insufficient memory to run query"}'
                    print("Input Json")
                    print(input_json)
                    driver.find_element(By.XPATH,f"{rs_text}").clear()
                    sleep(3)
                    driver.save_screenshot("dcsbv.png")
                    driver.find_element(By.XPATH,f"{rs_text}").send_keys(input_json)

                    #new try will try later starts

                    # try:
                    #     operation_div = WebDriverWait(driver, 15).until(
                    #     EC.visibility_of_element_located((By.ID, "operation_restartJobRun"))
                    # )

                    # # Now locate the visible textarea within the operation_div
                    #     textarea = operation_div.find_element(By.XPATH, "textarea[6]")  # Find textarea relative to the operation_div

                    #     # Ensure the textarea is interactable
                    #     if textarea.is_displayed() and textarea.is_enabled():
                    #         textarea.clear()  # Clear existing text
                    #         sleep(3)
                    #         driver.save_screenshot("dcsbv.png")
                    #         textarea.send_keys(input_json)  # Input the JSON

                    #         # Optionally take a screenshot for verification
                    #         driver.save_screenshot("textarea_filled.png")
                    #         print("Text area successfully filled with data.")
                    #     else:
                    #         print("The textarea is not interactable.")

                    # except Exception as e:
                    #     print(f"An error occurred while locating the textarea: {e}")
                    # sleep(3)
                    # driver.save_screenshot("dcsbv.png")
                    # # driver.find_element(By.XPATH,"/html/body/div[3]/div/div[103]/form/input").click()
                    # try:
                    #     submit_button = WebDriverWait(driver, 10).until(
                    #         EC.element_to_be_clickable((By.XPATH, "//input[@type='submit' and @value='Call restartJobRun']"))
                    #     )
                    #     submit_button.click()  # Click the submit button
                    # except Exception as e:
                    #     print(f"An error occurred: {e}")

                    #new try will try later ends
                    sleep(3)
                    driver.save_screenshot("dcsbv.png")
                    driver.find_element(By.XPATH,f"{rs_submit}").click()
                    sleep(3)
                    driver.save_screenshot("dcsbv.png")
                    # our = driver.find_element(By.CLASS_NAME,"spacer")
                    output_json = driver.find_element(By.XPATH,f"{rs_response}")
                    # print(output_json)
                    # print(our)
                    import json
                    print(output_json.text)
                    data = json.loads(output_json.text)

                    print(data['jobRunId'])
                    link_njrid = data['jobRunId']
                    link_njr = 'https://datacentral.a2z.com/console?action=jobrun_details&jobrun_id=' + str(link_njrid)
                    res_det = "Job has been restarted"
                    
                    # print(details)
                    table_rowserr['Sl No.'] = pd.Series([i] * 1)
                    table_rowserr['Table Name'] = pd.Series([name2] * 1)
                    table_rowserr['JobHistoryLink'] = pd.Series([link_jh] * 1)
                    table_rowserr['NewJobRunLink'] = pd.Series([link_njr] * 1)
                    table_rowserr['Dataset Date'] = pd.Series([dsdate] * 1)
                    table_rowserr['Status'] = pd.Series([stat] * 1)
                    table_rowserr['Details'] = pd.Series([res_det] * 1)
                    table_rowserr['Schedule Hour (Local)'] = pd.Series([schedule] * 1)
                    table_rowserr['SLA'] = pd.Series([res_sla] * 1)
                    table_rowserr['ElapsedHours'] = pd.Series([res_sla_elap] * 1)
                    table_rowserr['Expected End Time (IST)'] = pd.Series([res_et] * 1)
                    table_rowserr['SLA_FLAG'] = pd.Series([res_fl] * 1)
                    table_rowserr = table_rowserr[['Sl No.','Table Name','Status','Dataset Date','Schedule Hour (Local)','NewJobRunLink','JobHistoryLink','Details','SLA','ElapsedHours','Expected End Time (IST)','SLA_FLAG']]
                    print(table_rowserr)
                    print(table_rowserr.dtypes)
                    loop_restarted = pd.concat([table_rowserr],ignore_index=True)
                    df_restarted = df_restarted._append(loop_restarted,ignore_index=True)
                    print(df_restarted)
                except Exception as er:
                    excerr = pd.DataFrame()
                    name2 = row['Table Name']
                    stat = row['Status']
                    dsdate = row['Dataset Date']
                    schedule = row['Schedule Hour (Local)']
                    link_jr = row['JobRunLink']
                    link_jh = row['JobHistoryLink']
                    res_sla = row['SLA']
                    res_sla_elap = row['ElapsedHours']
                    res_et = row['Expected End Time (IST)']
                    res_fl = row['SLA_FLAG']
                    res_det = "Job could not be restarted"

                    excerr['Sl No'] = pd.Series([i] * 1)
                    excerr['Table Name'] = pd.Series([name2] * 1)
                    excerr['JobHistoryLink'] = pd.Series([link_jh] * 1)
                    excerr['JobRunLink'] = pd.Series([link_jr] * 1)
                    excerr['Dataset Date'] = pd.Series([dsdate] * 1)
                    excerr['Status'] = pd.Series([stat] * 1)
                    excerr['Details'] = pd.Series([res_det] * 1)
                    excerr['Schedule Hour (Local)'] = pd.Series([schedule] * 1)
                    excerr['SLA'] = pd.Series([res_sla] * 1)
                    excerr['ElapsedHours'] = pd.Series([res_sla_elap] * 1)
                    excerr['Expected End Time (IST)'] = pd.Series([res_et] * 1)
                    excerr['SLA_FLAG'] = pd.Series([res_fl] * 1)
                    excerr['Details'] = pd.Series([res_det] * 1)
                    excerr['Exception_Details'] =  pd.Series([str(er)] * 1) 
                    print(excerr)
                    loop1_exc2_ferr = pd.concat([excerr],ignore_index=True)
                    loop1_exc2err = loop1_exc2err._append(loop1_exc2_ferr,ignore_index=True)
            
            loop1_exc2err.to_excel("exception_in_restart.xlsx")
            df_restarted.to_excel("restartedjobs.xlsx") 
            if not loop1_exc2err.empty:
                loop1_exc2err['JobRunLink'] = loop1_exc2err['JobRunLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
                loop1_exc2err['JobHistoryLink'] = loop1_exc2err['JobHistoryLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
            print("Printing restart after loop")
            print(df_restarted)
            print("Printing exception in restart after loop")
            print(loop1_exc2err)
            if not df_restarted.empty:
                df_restarted['NewJobRunLink'] = df_restarted['NewJobRunLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
                df_restarted['JobHistoryLink'] = df_restarted['JobHistoryLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
            def highlightserr2(rows_er2):
                if rows_er2['Status']=='Error':
                    return ['background-color: red; color: white;'] * len(rows_er2)
                elif rows_er2['Status']=='Waiting for Dependencies' or rows_er2['Status']=='Waiting for Resources':
                    return ['background-color: linen; color: black;'] * len(rows_er2) #lavender
                elif rows_er2['Status']=='Executing':
                    return ['background-color: orange; color: black;'] * len(rows_er2)
                else:
                    return [''] * len(rows_er2)
            res = df_restarted.style.apply(highlightserr2, axis=1)
            table_style3 = [
            {'selector': 'table',
            'props': [
                ('border-collapse', 'collapse')
                #('border-radius', '500px')  # Adjust the radius as needed
            ]},
            {'selector': 'th, td',
            'props': [
                ('border', '1px solid black'),
                ('padding', '8px')
                #('border-radius', '500px')
            ]},
            {'selector': 'th',
            'props': [
                ('background-color', 'purple') #bisque
            ]}]
            res = res.set_table_styles(table_style3)
            res = res.set_table_attributes('style="border: 1px solid black;"')
            res.hide()
            bodyerrres = res.to_html()

            def highlightserr2x(rows_er2x):
                if rows_er2x['Status']=='Error':
                    return ['background-color: red; color: white;'] * len(rows_er2x)
                elif rows_er2x['Status']=='Waiting for Dependencies' or rows_er2x['Status']=='Waiting for Resources':
                    return ['background-color: linen; color: black;'] * len(rows_er2x) #lavender
                elif rows_er2x['Status']=='Executing':
                    return ['background-color: orange; color: black;'] * len(rows_er2x)
                else:
                    return [''] * len(rows_er2x)
            erx = loop1_exc2err.style.apply(highlightserr2x, axis=1)
            table_style4 = [
            {'selector': 'table',
            'props': [
                ('border-collapse', 'collapse')
                #('border-radius', '500px')  # Adjust the radius as needed
            ]},
            {'selector': 'th, td',
            'props': [
                ('border', '1px solid black'),
                ('padding', '8px')
                #('border-radius', '500px')
            ]},
            {'selector': 'th',
            'props': [
                ('background-color', 'purple') #bisque
            ]}]
            erx = erx.set_table_styles(table_style4)
            erx = erx.set_table_attributes('style="border: 1px solid black;"')
            erx.hide()
            bodyerrx = erx.to_html()

            #Sending mail for error and restart
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            #Outlook setup starts
            if teamss != '111':
                if teamss == '1':
                        # mail.To = 'kumraov'+'@amazon.com'
                    # mail.To = 'mdsagkha'+'@amazon.com;' + 'kumraov'+'@amazon.com;'
                    mail.To = 'inpay-dp-team@amazon.com;'  + 'mdsagkha'+'@amazon.com;'
                    mail.cc='sastry' + '@amazon.com;' +  'plammojo'+'@amazon.com;iankranj@amazon.com'
                
                else:
                    mail.cc = login + '@amazon.com;'

                mail.Body = 'Hi team,'+'\n\n'+'Please find the updated sheet for errored P0 tables (' + run + ' run).'+'\n'
                # mail.HTMLBody += "<b><font size='1'>Notes for SLA_FLAG:</font></b>"+'\n\n'+"<br><i><font size='1'>Breached: Breached SLA with any jobrun status (success, error, wfd & wfr)</font></i>"
                # mail.HTMLBody += "<br><i><font size='1'>Not Breached: Job run status is success and did not breach SLA</font></i>"+'\n'
                # mail.HTMLBody += "<br><i><font size='1'>Not Yet Breached: Job run status is not success and did not breach SLA</font></i><br>"+'\n'

                # mail.HTMLBody += "There were some code changes done, hence sending updated mail<br><br>"
                # mail.HTMLBody += "Sheet1 contains details from t-7 to t-4 and Sheet2 from t-3 to t-1<br><br>"
                mail.HTMLBody += "<b><i><u><font size='5'>All Errored Out Tables:</font></u></i></b><br>"
                # mail.HTMLBody += "<br><br>"
                mail.HTMLBody += bodyerr
                # mail.HTMLBody += "<br><br>"
                mail.HTMLBody += "<br><b><i><u><font size='5'>Restarted Tables Due To Reaper:</font></u></i></b><br>"
                # mail.HTMLBody += "<br><br>"
                mail.HTMLBody += bodyerrres
                mail.HTMLBody += "<br><b><i><u><font size='5'>Exception details during restart:</font></u></i></b><br>"
                # mail.HTMLBody += "<br><br>"
                mail.HTMLBody += bodyerrx
                mail.HTMLBody += "<br><b><font size='1'>Notes for SLA_FLAG:</font></b>"
                mail.HTMLBody +="<br><i><font size='1'>Breached: Breached SLA with any jobrun status (success, executing, error, wfd & wfr)</font></i>"
                mail.HTMLBody += "<br><i><font size='1'>Not Breached: Job run status is success and did not breach SLA</font></i>"
                mail.HTMLBody += "<br><i><font size='1'>Not Yet Breached: Job run status is not success and did not breach SLA</font></i><br>"
                mail.HTMLBody += f"<b><br>Thanks,<br>{mailname}</b>"
                #mail.Subject = 'Data Monitoring Sheet - Platform Team'
                mail.Subject = 'Errored!(' + run + ' run): Data Monitoring Sheet for P0 Tables - Platform Team'
                login = os.getlogin()
                attachment1  = "C:\\Users\\"+login+"\\Downloads\\dn_proj_reg\\"+"All_error_jobs.xlsx"
                attachment2  = "C:\\Users\\"+login+"\\Downloads\\dn_proj_reg\\"+"restartedjobs.xlsx"
                attachment3  = "C:\\Users\\"+login+"\\Downloads\\dn_proj_reg\\"+"exception_in_restart.xlsx"
                mail.Attachments.Add(attachment1)
                mail.Attachments.Add(attachment2)
                mail.Attachments.Add(attachment3)
                mail.Send()
                print('mail gaya')
                #Outlook setup ends
            else:
                        #Sending mail for error and restart
                outlook = win32.Dispatch('outlook.application')
                mail = outlook.CreateItem(0)
                #Outlook setup starts
                if teamss != '111':
                    if teamss == '1':
                        # mail.To = 'kumraov'+'@amazon.com'
                        # mail.To = 'mdsagkha'+'@amazon.com;' + 'kumraov'+'@amazon.com;'
                        mail.To = 'inpay-dp-team@amazon.com;'  + 'mdsagkha'+'@amazon.com;'
                        mail.cc='sastry' + '@amazon.com;' + 'plammojo'+'@amazon.com;iankranj@amazon.com'
                    
                    else:
                        mail.cc = login + '@amazon.com;'

                    mail.Body = 'Hi team,'+'\n\n'+'Please find the updated sheet for errored P0 tables (' + run + ' run).'+'\n'
                    # mail.HTMLBody += "<b><font size='1'>Notes for SLA_FLAG:</font></b>"+'\n\n'+"<br><i><font size='1'>Breached: Breached SLA with any jobrun status (success, error, wfd & wfr)</font></i>"
                    # mail.HTMLBody += "<br><i><font size='1'>Not Breached: Job run status is success and did not breach SLA</font></i>"+'\n'
                    # mail.HTMLBody += "<br><i><font size='1'>Not Yet Breached: Job run status is not success and did not breach SLA</font></i><br>"+'\n'

                    # mail.HTMLBody += "There were some code changes done, hence sending updated mail<br><br>"
                    # mail.HTMLBody += "Sheet1 contains details from t-7 to t-4 and Sheet2 from t-3 to t-1<br><br>"
                    mail.HTMLBody += "<b><i><u><font size='5'>All Errored Out Tables:</font></u></i></b><br>"
                    # mail.HTMLBody += "<br><br>"
                    mail.HTMLBody += bodyerr
                    mail.HTMLBody += "<br><b><font size='1'>Notes for SLA_FLAG:</font></b>"
                    mail.HTMLBody +="<br><i><font size='1'>Breached: Breached SLA with any jobrun status (success, executing, error, wfd & wfr)</font></i>"
                    mail.HTMLBody += "<br><i><font size='1'>Not Breached: Job run status is success and did not breach SLA</font></i>"
                    mail.HTMLBody += "<br><i><font size='1'>Not Yet Breached: Job run status is not success and did not breach SLA</font></i><br>"
                    mail.HTMLBody += f"<b><br>Thanks,<br>{mailname}</b>"
                    #mail.Subject = 'Data Monitoring Sheet - Platform Team'
                    mail.Subject = 'Errored!(' + run + ' run): Data Monitoring Sheet for P0 Tables - Platform Team'
                    login = os.getlogin()
                    attachment1  = "C:\\Users\\"+login+"\\Downloads\\dn_proj_reg\\"+"All_error_jobs.xlsx"
                    mail.Attachments.Add(attachment1)
                    mail.Send()
                    print('mail gaya')
                    #Outlook setup ends

    



    # Display the result
    print("Printing source error again")
    print(df_error)
    print("Union of df2_app_f1_error and df2_app_f7_error:")
    print(df_error2)
    print("Printing restart")
    print(df_restarted)
    print("Printing restart exception")
    print(loop1_exc2err)

    df2_app_f7_xl = df2_app_f7_xl.sort_values(by=['Sl No.'], ascending=[True])
    df2_app_f1_xl = df2_app_f1_xl.sort_values(by=['Sl No.'], ascending=[True])
    df1_success = df1_success.sort_values(by=['Sl No.'], ascending=[True])

    df2_app_f7['JobRunLink'] = df2_app_f7['JobRunLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
    df2_app_f7['JobHistoryLink'] = df2_app_f7['JobHistoryLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
    df2_app_f1['JobRunLink'] = df2_app_f1['JobRunLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
    df2_app_f1['JobHistoryLink'] = df2_app_f1['JobHistoryLink'].apply(lambda x: f'<a href="{x}">Click here</a>')

    df2_app_f7 = df2_app_f7.sort_values(by=['Sl No.'], ascending=[True])
    df2_app_f1 = df2_app_f1.sort_values(by=['Sl No.'], ascending=[True])

    df2_app_t12['JobRunLink'] = df2_app_t12['JobRunLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
    df2_app_t12['JobHistoryLink'] = df2_app_t12['JobHistoryLink'].apply(lambda x: f'<a href="{x}">Click here</a>')
    df2_app_t12 = df2_app_t12.sort_values(by=['Sl No.'], ascending=[True])
    df2_app_t12_mail = df2_app_t12
    if os.path.exists("SLA12_P0jobs_"+str(filedate)+".xlsx"):
        df2_app_t12_prev= pd.read_excel("SLA12_P0jobs_"+str(filedate)+".xlsx")
        df2_app_t12_prev_success = df2_app_t12_prev[df2_app_t12_prev['Status']=='Success']
        df2_app_t12_prev_success = df2_app_t12_prev_success[['Table Name','Dataset Date','SLA']]
        df2_app_t12_prev_success.rename(columns={'Table Name':'Table_Name'},inplace=True)
        df2_app_t12_prev_success.rename(columns={'Dataset Date':'Dataset_Date'},inplace=True)
        df2_app_t12_prev_success.rename(columns={'SLA':'SLA_'},inplace=True)

        total_prev_success = df2_app_t12_prev_success.shape[0]
        print(total_prev_success)
        print(df2_app_t12)
        print(df2_app_t12_prev_success)
        df2_app_t12_comp = pd.merge(df2_app_t12,df2_app_t12_prev_success,how='left',left_on=['Table Name','Dataset Date'],right_on=['Table_Name','Dataset_Date']).reset_index()
        df2_app_t12_comp_diff = df2_app_t12_comp[df2_app_t12_comp['SLA_'].isna()]
        print(df2_app_t12_comp_diff)
        df2_app_t12 = df2_app_t12_comp_diff
        df12_message = "There are total " + str(total_prev_success) + " successful jobruns available in previous run mail for SLA12_P0jobs"
    else:
        df2_app_t12 = df2_app_t12
        df12_message = "There are total 0 successful jobruns available in previous run mail for SLA12_P0jobs"
    df2_app_t12 = df2_app_t12[['Sl No.','Table Name','Status','Dataset Date','Schedule Hour (Local)','JobRunLink','JobHistoryLink','Sat_Dependency','Details','SLA','ElapsedHours','Expected End Time (IST)','Actual End Time (IST)','SLA_FLAG']]

    print(df2_app_t12)
    print(df12_message)

    df2_app_t12_mail.to_excel("SLA12_P0jobs_"+str(filedate)+".xlsx")    
    df2_app_t12_mail.to_excel("SLA12_P0jobs_"+str(filedate)+"_"+str(run)+"Run.xlsx")
    # def highlight_sla(val):
    #     if val == 'Breached':
    #         return 'background-color: maroon; color: white'
    #     return ''
    # def highlight_sla_cell(df):
    #     mask = df['SLA_FLAG'] == 'Breached'
    #     return pd.DataFrame('background-color: maroon; color: white;', 
    #                     index=df.index, 
    #                     columns=df.columns
    #                     ).where(mask & (df.columns == 'SLA_FLAG'), '')
    # def highlights(rows):
    #     if rows['Status']=='Error':
    #         return ['background-color: red; color: white;'] * len(rows)
    #     elif rows['Status']=='Waiting for Resources':
    #         return ['background-color: cornsilk; color: black;'] * len(rows)
    #     elif rows['Status']=='Waiting for Dependencies':
    #         return ['background-color: linen; color: black;'] * len(rows) #lavender
    #     elif rows['Status']=='Executing':
    #         return ['background-color: orange; color: black;'] * len(rows)
    #     else:
    #         return [''] * len(rows)
    def highlights(row):
        # Default style for the whole row based on the 'Status'
        row_style = [''] * len(row)  # Initialize with no styling
        
        if row['Status'] == 'Error':
            row_style = ['background-color: red; color: white;'] * len(row)
        elif row['Status'] == 'Waiting for Resources':
            row_style = ['background-color: cornsilk; color: black;'] * len(row)
        elif row['Status'] == 'Waiting for Dependencies':
            row_style = ['background-color: linen; color: black;'] * len(row)
        elif row['Status'] == 'Executing':
            row_style = ['background-color: orange; color: black;'] * len(row)
        
        # Now check and modify for sla_flag if it's "Breached"
        if row['SLA_FLAG'] == 'Breached':
            sla_flag_index = row.index.get_loc('SLA_FLAG')  # Find the index of the 'sla_flag' column
            row_style[sla_flag_index] = 'background-color: maroon; color: white;'  # Highlight the cell
        
        return row_style   
    r = df2_app_f7.style.apply(highlights, axis=1)
    # r = df2_app_f7.style\
    # .apply(highlights, axis=1)\
    # .apply(highlight_sla, subset=['SLA_FLAG'])
    #r = df2_app_f7.style.apply(highlight_sla_cell, axis=None)

    # r_xl = df2_app_f7_xl.style.apply(highlights, axis=1)

    table_style = [
        {'selector': 'table',
        'props': [
            ('border-collapse', 'collapse')
            #('border-radius', '500px')  # Adjust the radius as needed
        ]},
        {'selector': 'th, td',
        'props': [
            ('border', '1px solid black'),
            ('padding', '8px'),
            ('font-size', '8px')
            #('border-radius', '500px')
        ]},
        {'selector': 'th',
        'props': [
            ('background-color', 'bisque') #purple
            ,('font-size', '9px')
        ]}
    ]

    r = r.set_table_styles(table_style)
    r = r.set_table_attributes('style="border: 1px solid black;"')
    r.hide()
    # r_xl = r_xl.set_table_styles(table_style)
    # r_xl = r_xl.set_table_attributes('style="border: 1px solid black;"')
    # r_xl.hide()
    body1 = r.to_html()

    # def highlight(row):
    #     if row['Status']=='Error':
    #         return ['background-color: red; color: white;'] * len(row)
    #     elif row['Status']=='Waiting for Resources':
    #         return ['background-color: cornsilk; color: black;'] * len(row)
    #     elif row['Status']=='Waiting for Dependencies':
    #         return ['background-color: lavender; color: black;'] * len(row) #linen
    #     elif row['Status']=='Executing':
    #         return ['background-color: orange; color: black;'] * len(row)
    #     else:
    #         return [''] * len(row)
    def highlight(row):
        # Default style for the whole row based on the 'Status'
        row_style = [''] * len(row)  # Initialize with no styling
        
        # Set the row style based on the Status
        if row['Status'] == 'Error':
            row_style = ['background-color: red; color: white;'] * len(row)
        elif row['Status'] == 'Waiting for Resources':
            row_style = ['background-color: cornsilk; color: black;'] * len(row)
        elif row['Status'] == 'Waiting for Dependencies':
            row_style = ['background-color: lavender; color: black;'] * len(row)
        elif row['Status'] == 'Executing':
            row_style = ['background-color: orange; color: black;'] * len(row)
        
        # Now check if sla_flag is "Breached" and highlight only that cell
        if row['SLA_FLAG'] == 'Breached':
            sla_flag_index = row.index.get_loc('SLA_FLAG')  # Find the index of the 'sla_flag' column
            row_style[sla_flag_index] = 'background-color: maroon; color: white;'  # Highlight the sla_flag cell
        
        return row_style
    # styled_df = df.style\
    # .apply(highlight_status, axis=1)\
    # .apply(highlight_sla, subset=['SLA_FLAG'])
    s = df2_app_f1.style.apply(highlight, axis=1)
    # s = df2_app_f1.style\
    # .apply(highlight, axis=1)\
    # .apply(highlight_sla, subset=['SLA_FLAG'])
    # s_xl = df2_app_f1_xl.style.apply(highlight, axis=1)

    table_style2 = [
        {'selector': 'table',
        'props': [
            ('border-collapse', 'collapse')
            #('border-radius', '500px')  # Adjust the radius as needed
        ]},
        {'selector': 'th, td',
        'props': [
            ('border', '1px solid black'),
            ('padding', '8px'),
            ('font-size', '8px')
            #('border-radius', '500px')
        ]},
        {'selector': 'th',
        'props': [
            ('background-color', 'purple') #bisque
            ,('font-size', '9px')
        ]}
    ]

    table_style_t12 = [
        {'selector': 'table',
        'props': [
            ('border-collapse', 'collapse')
            #('border-radius', '500px')  # Adjust the radius as needed
        ]},
        {'selector': 'th, td',
        'props': [
            ('border', '1px solid black'),
            ('padding', '8px'),
            ('font-size', '8px')
            #('border-radius', '500px')
        ]},
        {'selector': 'th',
        'props': [
            ('background-color', 'lightgreen') #purple
            ,('font-size', '9px')
        ]}
    ]
    s = s.set_table_styles(table_style2)
    s = s.set_table_attributes('style="border: 1px solid black;"')
    s.hide()
    # s_xl = s_xl.set_table_styles(table_style2)
    # s_xl = s_xl.set_table_attributes('style="border: 1px solid black;"')
    # s_xl.hide()
    body2 = s.to_html()


    t12 = df2_app_t12.style.apply(highlight, axis=1)
    # t12 = df2_app_t12.style\
    # .apply(highlight, axis=1)\
    # .apply(highlight_sla, subset=['SLA_FLAG'])
    t12 = t12.set_table_styles(table_style_t12)
    t12 = t12.set_table_attributes('style="border: 1px solid black;"')
    t12.hide()

    body_t12 = t12.to_html()


    print(df1_app)
    df1_app.to_excel("Checkmari.xlsx")

    print(df2_app_f7)
    print(df2_app_f1)
    print(r)
    print(s)


    # The manipulator
    orange = "FFA500"
    purple = "800080"
    whitesmoke = "F5F5F5"
    beige = "F5F5DC"
    lavender = "E6E6FA"
    bisque = "FFE4C4"
    linen = "FAF0E6"
    lightgreen = "90EE90"
    lightblue = "ADD8E6"
    red = "FF0000"
    green = "008000"
    blue = "0000FF"
    RED = "FFFF0000"  # Red
    GREEN = "FF00FF00"  # Green
    BLUE = "FF0000FF"  # Blue

    # Success Table
    wb = Workbook()
    ws = wb.active
    for sc in dataframe_to_rows(df1_success, index=False, header=True):
        ws.append(sc)
    color_dict = {
        "Success": GREEN  # Red
        # Green
    }
    hd = BLUE
    for row in range(2, len(df1_success) + 2):
        status = ws.cell(row, 3).value  # Assuming the 'country' column is in column C
        if status in color_dict:
            fill = PatternFill(start_color=color_dict[status], end_color=color_dict[status], fill_type="solid")
            for cell in ws[row]:
                cell.fill = fill
    for cell in ws[1]:
        cell.fill = PatternFill(start_color=hd, end_color=hd, fill_type="solid")
    for index, row in df1_success.iterrows():
        cell1 = f'F{index+2}'  # Assuming the 'link' column is in column B
        link1 = row['JobRunLink']
        ws[cell1].hyperlink = Hyperlink(ref=link1, target=link1)
        ws[cell1].value = 'Click Here'  # Set the display text to 'Click Here'
        cell2 = f'G{index+2}'  # Assuming the 'link' column is in column B
        link2 = row['JobHistoryLink']
        ws[cell2].hyperlink = Hyperlink(ref=link2, target=link2)
        ws[cell2].value = 'Click Here'  # Set the display text to 'Click Here'

    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, min_col=1, max_row=len(df1_success)+1, max_col=len(df1_success.columns)):
        for cell in row:
            cell.border = border

    excel_file_path = "Success_T-1_P0jobs_"+str(filedate)+"_"+str(run)+"Run.xlsx"
    wb.save(excel_file_path)

    # T-2 to T-7 Table
    wb = Workbook()
    ws = wb.active
    for sc in dataframe_to_rows(df2_app_f7_xl, index=False, header=True):
        ws.append(sc)
    color_dict = {
        "Error": red,
        "Executing": orange,
        "Waiting for Resources": linen,
        "Waiting for Dependencies": linen  # Red
        # Green
    }
    hd = bisque
    for row in range(2, len(df2_app_f7_xl) + 2):
        status = ws.cell(row, 3).value  # Assuming the 'country' column is in column C
        if status in color_dict:
            fill = PatternFill(start_color=color_dict[status], end_color=color_dict[status], fill_type="solid")
            for cell in ws[row]:
                cell.fill = fill
    for cell in ws[1]:
        cell.fill = PatternFill(start_color=hd, end_color=hd, fill_type="solid")
    for index, row in df2_app_f7_xl.iterrows():
        cell1 = f'F{index+2}'  # Assuming the 'link' column is in column B
        link1 = row['JobRunLink']
        ws[cell1].hyperlink = Hyperlink(ref=link1, target=link1)
        ws[cell1].value = 'Click Here'  # Set the display text to 'Click Here'
        cell2 = f'G{index+2}'  # Assuming the 'link' column is in column B
        link2 = row['JobHistoryLink']
        ws[cell2].hyperlink = Hyperlink(ref=link2, target=link2)
        ws[cell2].value = 'Click Here'  # Set the display text to 'Click Here'

    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, min_col=1, max_row=len(df2_app_f7_xl)+1, max_col=len(df2_app_f7_xl.columns)):
        for cell in row:
            cell.border = border

    excel_file_path = "P0jobs_Delayed_Tables_"+str(filedate)+"_"+str(run)+"Run.xlsx"
    wb.save(excel_file_path)

    # T-1 Table
    wb = Workbook()
    ws = wb.active
    for sc in dataframe_to_rows(df2_app_f1_xl, index=False, header=True):
        ws.append(sc)
    color_dict = {
        "Error": red,
        "Executing": orange,
        "Waiting for Resources": lavender,
        "Waiting for Dependencies": lavender
    }
    hd = purple
    for row in range(2, len(df2_app_f1_xl) + 2):
        status = ws.cell(row, 3).value  # Assuming the 'country' column is in column C
        if status in color_dict:
            fill = PatternFill(start_color=color_dict[status], end_color=color_dict[status], fill_type="solid")
            for cell in ws[row]:
                cell.fill = fill
    for cell in ws[1]:
        cell.fill = PatternFill(start_color=hd, end_color=hd, fill_type="solid")
    for index, row in df2_app_f1_xl.iterrows():
        cell1 = f'F{index+2}'  # Assuming the 'link' column is in column B
        link1 = row['JobRunLink']
        ws[cell1].hyperlink = Hyperlink(ref=link1, target=link1)
        ws[cell1].value = 'Click Here'  # Set the display text to 'Click Here'
        cell2 = f'G{index+2}'  # Assuming the 'link' column is in column B
        link2 = row['JobHistoryLink']
        ws[cell2].hyperlink = Hyperlink(ref=link2, target=link2)
        ws[cell2].value = 'Click Here'  # Set the display text to 'Click Here'

    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, min_col=1, max_row=len(df2_app_f1_xl)+1, max_col=len(df2_app_f1_xl.columns)):
        for cell in row:
            cell.border = border

    excel_file_path = "P0jobs_T-1_"+str(filedate)+"_"+str(run)+"Run.xlsx"
    wb.save(excel_file_path)
    # End saving excels
    # Load the first Excel file
    # r_xl.to_excel("Regularjobs_T-2toT-7_"+str(filedate)+"_"+str(run)+"Run.xlsx",index=False)
    # s_xl.to_excel("Regularjobs_T-1_"+str(filedate)+"_"+str(run)+"Run.xlsx",index=False)
    # df1_success.to_excel("Success_T-1_Regularjobs_"+str(filedate)+"_"+str(run)+"Run.xlsx",index=False)

    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    #Outlook setup starts
    if teamss != '111':
        if teamss == '1':
                        # mail.To = 'kumraov'+'@amazon.com'
                        # mail.To = 'mdsagkha'+'@amazon.com;' + 'kumraov'+'@amazon.com;'
            mail.To = 'inpay-dp-team@amazon.com;'  + 'mdsagkha'+'@amazon.com;'
            mail.cc='sastry' + '@amazon.com;'  + 'plammojo'+'@amazon.com;iankranj@amazon.com'
        
        else:
            mail.cc = login + '@amazon.com;'

        mail.Body = 'Hi team,'+'\n\n'+'Please find the updated sheet for data monitoring job for P0 tables (' + run + ' run).'+'\n'
        # mail.HTMLBody += "<b><font size='1'>Notes for SLA_FLAG:</font></b>"+'\n\n'+"<br><i><font size='1'>Breached: Breached SLA with any jobrun status (success, error, wfd & wfr)</font></i>"
        # mail.HTMLBody += "<br><i><font size='1'>Not Breached: Job run status is success and did not breach SLA</font></i>"+'\n'
        # mail.HTMLBody += "<br><i><font size='1'>Not Yet Breached: Job run status is not success and did not breach SLA</font></i><br>"+'\n'

        # mail.HTMLBody += "There were some code changes done, hence sending updated mail<br><br>"
        # mail.HTMLBody += "Sheet1 contains details from t-7 to t-4 and Sheet2 from t-3 to t-1<br><br>"

        mail.HTMLBody += "<b><i><u><font size='3'>SLA-12 Tables (with all statuses and upto T-7 dataste date):</font></u></i></b><br>"
        # mail.HTMLBody += "<br><br>"
        mail.HTMLBody += f"<b><i>{df12_message}</i></b><br>"
        mail.HTMLBody += "<b><i>For further details check the attachment for SLA12_P0jobs in this mail (current run).</i></b><br>"

        mail.HTMLBody += body_t12
        mail.HTMLBody += "<b><i><u><font size='3'>Delayed Tables (Might breached SLA):</font></u></i></b><br>"
        # mail.HTMLBody += "<br><br>"
        mail.HTMLBody += body1
        # mail.HTMLBody += "<br><br>"
        mail.HTMLBody += "<br><b><i><u><font size='3'>T-1:</font></u></i></b><br>"
        # mail.HTMLBody += "<br><br>"
        mail.HTMLBody += body2
        mail.HTMLBody += "<br><b><font size='1'>Notes for SLA_FLAG:</font></b>"
        mail.HTMLBody +="<br><i><font size='1'>Breached: Breached SLA with any jobrun status (success, executing, error, wfd & wfr)</font></i>"
        mail.HTMLBody += "<br><i><font size='1'>Not Breached: Job run status is success and did not breach SLA</font></i>"
        mail.HTMLBody += "<br><i><font size='1'>Not Yet Breached: Job run status is not success and did not breach SLA</font></i><br>"
        mail.HTMLBody += f"<b><br>Thanks,<br>{mailname}</b>"
        #mail.Subject = 'Data Monitoring Sheet - Platform Team'
        mail.Subject = '(' + run + ' run): Data Monitoring Sheet for P0 Tables - Platform Team'
        login = os.getlogin()
        attachment1  = "C:\\Users\\"+login+"\\Downloads\\dn_proj_reg\\"+"P0jobs_Delayed_Tables_"+str(filedate)+"_"+str(run)+"Run.xlsx"
        attachment2  = "C:\\Users\\"+login+"\\Downloads\\dn_proj_reg\\"+"P0jobs_T-1_"+str(filedate)+"_"+str(run)+"Run.xlsx"
        attachment3  = "C:\\Users\\"+login+"\\Downloads\\dn_proj_reg\\"+"Success_T-1_P0jobs_"+str(filedate)+"_"+str(run)+"Run.xlsx"
        attachment4  = "C:\\Users\\"+login+"\\Downloads\\dn_proj_reg\\"+"SLA12_P0jobs_"+str(filedate)+"_"+str(run)+"Run.xlsx"
        mail.Attachments.Add(attachment1)
        mail.Attachments.Add(attachment2)
        mail.Attachments.Add(attachment3)
        mail.Attachments.Add(attachment4)

        mail.Send()
        print('mail gaya')

        #Outlook setup ends
    current_timestamp2 = time.time()
    endtime = current_timestamp2
    print(endtime)
    dt_object2 = datetime.datetime.fromtimestamp(endtime)
    print("End time")
    print(dt_object2)
    formatted2 = dt_object2.strftime("%Y-%m-%d %H:%M:%S")
    print(formatted2)
    print("Time taken for excecution")
    difference = endtime - starttime
    timedelta = datetime.timedelta(seconds=difference)
    print(timedelta)
    # Write the updated count back to the file
    with open(count_file_name, "w") as count_file:
        count_file.write(str(run_count))
    # Sleep for 3 hours
    time.sleep(3 * 60 * 60)  # Sleep for 3 hours
ctypes.windll.kernel32.SetThreadExecutionState(0x80000000)
